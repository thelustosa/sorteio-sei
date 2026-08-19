#!/usr/bin/env python3
"""Converte a planilha da Câmara de Julgamento em SQL de importação.

    python dados/importar_planilha.py "Câmara de Julgamento - REG.xlsx"

Gera dados/acervo_cj.sql e dados/julgados_cj.sql. Rode os dois, nessa ordem, no
SQL Editor do Supabase — o acervo primeiro, senão os julgados não encontram o
processo para se ligar.

Os arquivos são idempotentes (ON CONFLICT DO NOTHING): rodar de novo não
duplica nada. Eles NÃO entram no git (ver .gitignore): carregam nome de
interessado pessoa física e o repositório é público.

Da aba Julgados são lidos os VALORES já calculados pelas fórmulas, não as
fórmulas: relator, defesa e data de distribuição entram como o histórico
registrou. O gatilho do banco só preenche o que vier em branco, então nada é
reescrito na importação — ele apenas amarra cada julgado ao registro do acervo.
"""

import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path

import openpyxl

# Rótulos que a CJ usa; a planilha tem "VISTA" e "Vista" na mesma coluna e isso
# viraria duas categorias em qualquer relatório. Valor fora da lista passa
# intacto — não é papel da importação inventar rótulo novo.
VOTOS = ['Manter', 'Anular', 'Vista']
STATUS = ['Julgado', 'Retornou', 'Retirado', 'Vista']

LOTE = 500  # linhas por comando INSERT


def texto(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def rotulo(v, conhecidos):
    s = texto(v)
    if s is None:
        return None
    for c in conhecidos:
        if s.casefold() == c.casefold():
            return c
    return s


def processo(v):
    """Número SEI só com dígitos: é a chave de ligação entre acervo e julgados."""
    s = texto(v)
    if s is None:
        return None
    d = ''.join(ch for ch in s if ch.isdigit())
    return d or None


def defesa(v):
    s = texto(v)
    if s is None:
        return None
    s = unicodedata.normalize('NFKD', s.casefold())
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return {'sim': True, 'nao': False}.get(s)


def dia(v):
    if isinstance(v, datetime):
        return v.date()
    return v if isinstance(v, date) else None


def inteiro(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def literal(v):
    if v is None:
        return 'null'
    if v is True:
        return 'true'
    if v is False:
        return 'false'
    if isinstance(v, int):
        return str(v)
    if isinstance(v, date):
        return f"date '{v.isoformat()}'"
    return "'" + str(v).replace("'", "''") + "'"


def ler_acervo(ws):
    linhas = []
    for r in range(2, ws.max_row + 1):
        num = processo(ws.cell(r, 1).value)
        if num is None:
            continue  # a aba termina com linhas em branco formatadas
        linhas.append({
            'num_processo': num,
            'relator': texto(ws.cell(r, 2).value),
            'data_distribuicao': dia(ws.cell(r, 3).value),
            'defesa': defesa(ws.cell(r, 4).value),
            'origem': 'planilha',
        })
    return linhas


def ler_julgados(ws):
    linhas = []
    for r in range(2, ws.max_row + 1):
        num = processo(ws.cell(r, 1).value)
        if num is None:
            continue
        linhas.append({
            'num_processo': num,
            'interessado': texto(ws.cell(r, 2).value),
            'defesa': defesa(ws.cell(r, 3).value),
            'data_sessao': dia(ws.cell(r, 4).value),
            'voto': rotulo(ws.cell(r, 5).value, VOTOS),
            'pauta': inteiro(ws.cell(r, 6).value),
            'status': rotulo(ws.cell(r, 7).value, STATUS),
            'data_distribuicao': dia(ws.cell(r, 9).value),
            'relator': texto(ws.cell(r, 11).value),
        })
    return linhas


def deduplicar(linhas, chave):
    """Tira as cópias exatas de digitação, preservando a ordem da planilha."""
    vistos, unicas, repetidas = set(), [], 0
    for l in linhas:
        k = tuple(l[c] for c in chave)
        if k in vistos:
            repetidas += 1
            continue
        vistos.add(k)
        unicas.append(l)
    return unicas, repetidas


def gerar_sql(tabela, colunas, linhas, restricao, cabecalho):
    partes = [f'-- {cabecalho}\n-- Gerado por dados/importar_planilha.py — não editar à mão.\n']
    for i in range(0, len(linhas), LOTE):
        lote = linhas[i:i + LOTE]
        valores = ',\n  '.join(
            '(' + ', '.join(literal(l.get(c)) for c in colunas) + ')' for l in lote
        )
        partes.append(
            f"insert into public.{tabela} ({', '.join(colunas)}) values\n  {valores}\n"
            f'on conflict on constraint {restricao} do nothing;\n'
        )
    return '\n'.join(partes)


def main(argv):
    if len(argv) != 2:
        print(__doc__)
        return 2

    planilha = Path(argv[1])
    if not planilha.is_file():
        print(f'Planilha não encontrada: {planilha}')
        return 1

    saida = Path(__file__).resolve().parent
    wb = openpyxl.load_workbook(planilha, data_only=True)

    acervo, repet_a = deduplicar(
        ler_acervo(wb['Acervo']), ('num_processo', 'data_distribuicao', 'relator'))
    julgados, repet_j = deduplicar(
        ler_julgados(wb['Julgados']), ('num_processo', 'data_sessao'))
    wb.close()

    (saida / 'acervo_cj.sql').write_text(gerar_sql(
        'acervo_cj',
        ['num_processo', 'relator', 'data_distribuicao', 'defesa', 'origem'],
        acervo, 'acervo_cj_distribuicao_unica',
        f'Acervo da Câmara de Julgamento — {len(acervo)} distribuições ({planilha.name}).',
    ), encoding='utf-8')

    (saida / 'julgados_cj.sql').write_text(gerar_sql(
        'julgados_cj',
        ['num_processo', 'interessado', 'data_sessao', 'pauta', 'voto', 'status',
         'defesa', 'relator', 'data_distribuicao'],
        julgados, 'julgados_cj_sessao_unica',
        f'Julgados da Câmara de Julgamento — {len(julgados)} sessões ({planilha.name}).',
    ), encoding='utf-8')

    print(f'acervo_cj.sql   {len(acervo):5d} linhas  ({repet_a} cópias exatas descartadas)')
    print(f'julgados_cj.sql {len(julgados):5d} linhas  ({repet_j} cópias exatas descartadas)')
    print(f'Rode acervo_cj.sql antes de julgados_cj.sql. Pasta: {saida}')
    return 0


if __name__ == '__main__':
    sys.exit(main(sys.argv))
