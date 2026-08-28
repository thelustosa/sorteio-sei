#!/usr/bin/env python3
"""Converte as planilhas do Conselho Regulador em SQL de importação.

    python dados/importar_creg.py "\\\\srvti-agr\\Repositorios_PowerBI\\Conselho Regulador"

Gera dados/acervo_creg.sql e dados/julgados_creg.sql. Rode os dois, nessa ordem,
no SQL Editor do Supabase — o acervo primeiro, senão os julgados não encontram o
processo para se ligar.

Os arquivos são idempotentes (ON CONFLICT DO NOTHING): rodar de novo não
duplica nada. Eles NÃO entram no git (ver .gitignore).

São seis planilhas, com dois papéis:

  CREG1.xlsx … CREG4.xlsx     o acervo, uma por gabinete. A unidade vem do
                              ARQUIVO, não da coluna "Gabinete" — é assim que a
                              planilha-mãe sempre decidiu (as fórmulas
                              procuravam o processo em [2]Planilha1, depois
                              [3], [4] e [5], e "Unidade CREG" recebia 1..4
                              conforme onde achasse), e a coluna tem ao menos
                              uma linha marcada CREG4 dentro do CREG3.

  Conselho Regulador.xlsx     os julgados de 2023 a maio/2025, em três abas
                              ("Página 2023", "Página 2024", "Página 2025").
                              As duas primeiras têm o cabeçalho na linha 3 e as
                              colunas deslocadas duas casas à direita.

  Conselho Regulador 2025.2.xlsx  os julgados de junho/2025 em diante, na aba
                              "Página 2025.2", já com o cabeçalho na linha 1.

Das abas Página são lidos os VALORES calculados pelas fórmulas, não as
fórmulas: assunto, recurso, unidade e data de distribuição entram como o
histórico registrou, e o gatilho do banco só preenche o que vier em branco.
"""

import collections
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from importar_planilha import LOTE, deduplicar, dia, inteiro, literal, texto  # noqa: E402

# As abas de julgados e onde cada uma guarda o cabeçalho e a primeira coluna.
# Em 2023 e 2024 a planilha tem duas colunas em branco à esquerda e um título
# ocupando as duas primeiras linhas; o resto do layout é idêntico.
PAGINAS = [
    ('Conselho Regulador.xlsx',        'Página 2023',   3, 3),
    ('Conselho Regulador.xlsx',        'Página 2024',   3, 3),
    ('Conselho Regulador.xlsx',        'Página 2025',   1, 1),
    ('Conselho Regulador 2025.2.xlsx', 'Página 2025.2', 1, 1),
]

# Rótulos que o Conselho usa hoje. O histórico tem 23 grafias de voto — quatro
# formas de dizer "aprovação", duas de "indeferimento", duas de "extinção" — e
# elas viravam categorias separadas em qualquer relatório.
#
# Só a GRAFIA é corrigida: "Aprovado" e "Apovação" são a mesma decisão escrita
# de outro jeito. Rótulo que não tem equivalente na lista ("Parcialmente
# Deferido", "Suspender", "Improvimento" — 13 linhas em 4.565) passa intacto,
# com a caixa arrumada: colapsá-lo num vizinho seria reescrever a decisão.
VOTOS = {
    'manter': 'Manter',
    'anular': 'Anular',
    'aprovacao': 'Aprovação', 'aprovado': 'Aprovação', 'apovacao': 'Aprovação',
    'indeferimento': 'Indeferimento', 'indeferir': 'Indeferimento',
    'extincao': 'Extinção', 'extinto': 'Extinção',
    'retirado': 'Retirado',
    'vista': 'Vista',
}
STATUS = {
    'julgado': 'Julgado', 'retirado': 'Retirado', 'vista': 'Vista',
    'sobrestado': 'Sobrestado', 'prejudicado': 'Prejudicado',
}

# Assunto só tem problema de CAIXA: a mesma planilha traz "AUTO DE INFRAÇÃO" e
# "Auto de Infração". Aqui a comparação ignora caixa e acento e devolve a forma
# canônica; o que não estiver na lista ("Revisão de Metodologia", "POP") passa
# como veio — inventar categoria seria pior do que ter uma a mais.
#
# A lista tem de acompanhar assuntosCreg, em assets/js/index.js: é ela que a
# secretaria vê no sorteio, e um rótulo só existente de um lado viraria duas
# categorias no relatório. "Quadro de Horários" entrou em 27/08/2026, e o
# histórico já trazia "Quadro de Horário" (singular) em 3 linhas — que a
# comparação sem caixa nem acento NÃO alcança, por serem palavras diferentes.
ASSUNTOS = ['Auto de Infração', 'Chamamento Público', 'Gratuidade',
            'Manifestação', 'Minuta', 'Nota Técnica', 'Ouvidoria',
            'Requerimento', 'Plano de Racionamento', 'Quadro de Horários',
            'Reajuste', 'Outros']

# A planilha escrevia "n/a" onde a fórmula não achava nada. No banco isso é
# ausência de dado, não um valor.
VAZIOS = {'n/a', 'n.a', 'na', '#n/d', '#n/a', '-'}

TIPOS = {
    'num_processo': 'text', 'unidade': 'text', 'data_distribuicao': 'date',
    'assunto': 'text', 'recurso': 'text', 'ordem': 'int', 'origem': 'text',
    'data_sessao': 'date', 'pauta': 'int', 'voto': 'text', 'status': 'text',
    'defesa': 'boolean', 'data_dist_cj': 'date', 'relator_cj': 'text',
    'voto_cj': 'text',
}


def chave(v):
    """Texto sem acento, minúsculo e sem espaço sobrando — só para comparar."""
    s = texto(v)
    if s is None:
        return None
    s = unicodedata.normalize('NFKD', ' '.join(s.split()).casefold())
    return ''.join(c for c in s if not unicodedata.combining(c))


def limpo(v):
    """Texto útil, ou None quando a planilha só tinha um marcador de vazio."""
    s = texto(v)
    return None if s is None or chave(s) in VAZIOS else ' '.join(s.split())


def rotulo(v, mapa):
    """A forma do Conselho, ou o texto como a planilha escreveu."""
    s = chave(v)
    if s is None or s in VAZIOS:
        return None
    return mapa.get(s) or limpo(v)


def assunto(v):
    s = chave(v)
    if s is None or s in VAZIOS:
        return None
    for a in ASSUNTOS:
        if chave(a) == s:
            return a
    return limpo(v)


def unificar(linhas, campo):
    """Uma grafia por rótulo: a mais usada entre as variantes da planilha.

    O mesmo assunto aparece escrito de vários jeitos, e cada grafia vira uma
    categoria em qualquer relatório. Title case resolveria a caixa e destruiria
    as siglas — "MINUTA RN" viraria "Minuta Rn", "Duplicidade AI" viraria
    "Duplicidade Ai", "(BPe)" viraria "(Bpe)". Escolher entre as grafias que a
    planilha de fato usou preserva as duas coisas.

    Desempate: frequência primeiro e, em seguida, a que tem menos caixa alta —
    entre "PLANEJAMENTO ESTRATÉGICO" e "Planejamento Estratégico", uma ocorrência
    cada, fica a segunda.
    """
    grafias = collections.defaultdict(collections.Counter)
    for l in linhas:
        if l.get(campo):
            grafias[chave(l[campo])][l[campo]] += 1

    preferida = {
        k: max(v, key=lambda t: (v[t], -sum(c.isupper() for c in t)))
        for k, v in grafias.items()
    }
    for l in linhas:
        if l.get(campo):
            l[campo] = preferida[chave(l[campo])]
    return linhas


def booleano(v):
    return {'sim': True, 'nao': False}.get(chave(v))


def processo(v):
    """Número SEI: 15 dígitos. É a chave de ligação entre acervo e julgados."""
    s = texto(v)
    if s is None:
        return None
    d = ''.join(c for c in s if c.isdigit())
    return d if len(d) == 15 else None


def unidade(v):
    """"1" ou 1 -> "CREG1". Qualquer outra coisa é vazio.

    A coluna "Unidade CREG" da planilha guarda o número do gabinete, mas 15
    linhas de 2024 têm anotação de sessão no lugar ("VISTA", "PC (VISTA)").
    Deixá-las nulas faz o gatilho derivar a unidade do acervo, que é onde o
    dado realmente está.
    """
    s = texto(v)
    if s is None:
        return None
    s = s.split('.')[0]
    return f'CREG{s}' if s.isdigit() and 1 <= int(s) <= 9 else None


def ler_acervo(pasta):
    """(distribuições, linhas descartadas) das quatro planilhas de gabinete."""
    linhas, descartadas = [], []
    for n in (1, 2, 3, 4):
        arquivo = pasta / f'CREG{n}.xlsx'
        wb = openpyxl.load_workbook(arquivo, data_only=True)
        ws = wb['Planilha1']
        for r in range(2, ws.max_row + 1):
            bruto = ws.cell(r, 2).value
            num = processo(bruto)
            data = dia(ws.cell(r, 5).value)
            # A aba termina com centenas de linhas em branco formatadas: linha
            # sem nada na coluna do processo não é perda, é o fim da tabela.
            if bruto is None and data is None:
                continue
            # Já estas são perda de verdade, e têm de aparecer no relatório: um
            # número fora dos 15 dígitos (erro de digitação) ou uma distribuição
            # sem data, que não entra num acervo que conta dias parados.
            if num is None or data is None:
                descartadas.append(
                    f'CREG{n}.xlsx linha {r}: '
                    + ('processo ' + repr(texto(bruto)) if num is None
                       else 'sem data de distribuição (processo ' + num + ')'))
                continue
            linhas.append({
                'num_processo': num,
                'unidade': f'CREG{n}',
                'data_distribuicao': data,
                'assunto': assunto(ws.cell(r, 4).value),
                'recurso': limpo(ws.cell(r, 6).value),
                'origem': 'planilha',
            })
        wb.close()
    return linhas, descartadas


def ler_julgados(pasta):
    """(julgados, linhas descartadas) das abas Página."""
    linhas, descartadas = [], []
    for arquivo, aba, cabecalho, primeira in PAGINAS:
        wb = openpyxl.load_workbook(pasta / arquivo, data_only=True)
        ws = wb[aba]
        desloca = primeira - 1
        col = lambda r, c: ws.cell(r, desloca + c).value  # noqa: E731
        for r in range(cabecalho + 1, ws.max_row + 1):
            bruto = col(r, 1)
            num = processo(bruto)
            sessao = dia(col(r, 20))
            if bruto is None and sessao is None:
                continue
            if num is None or sessao is None:
                descartadas.append(
                    f'{aba} linha {r}: '
                    + ('processo ' + repr(texto(bruto)) if num is None
                       else 'sem data de sessão (processo ' + num + ')'))
                continue
            linhas.append({
                'num_processo': num,
                'data_sessao': sessao,
                'pauta': inteiro(col(r, 19)),
                'voto': rotulo(col(r, 22), VOTOS),
                'status': rotulo(col(r, 21), STATUS),
                'unidade': unidade(col(r, 18)),
                'data_distribuicao': dia(col(r, 17)),
                'assunto': assunto(col(r, 3)),
                'recurso': limpo(col(r, 16)),
                'defesa': booleano(col(r, 12)),
                'data_dist_cj': dia(col(r, 13)),
                'relator_cj': limpo(col(r, 14)),
                'voto_cj': rotulo(col(r, 15), VOTOS),
            })
        wb.close()
    return linhas, descartadas


def gerar_sql(tabela, colunas, linhas, restricao, cabecalho):
    partes = [f'-- {cabecalho}\n-- Gerado por dados/importar_creg.py — não editar à mão.\n']
    lista = ', '.join(colunas)
    for i in range(0, len(linhas), LOTE):
        lote = linhas[i:i + LOTE]
        valores = ',\n    '.join(
            '(' + ', '.join(
                literal(l.get(c)) + (f'::{TIPOS[c]}' if n == 0 else '') for c in colunas
            ) + ')'
            for n, l in enumerate(lote)
        )
        partes.append(
            f'insert into public.{tabela} ({lista})\nvalues\n    {valores}\n'
            f'on conflict on constraint {restricao} do nothing;\n'
        )
    return '\n'.join(partes)


def main(argv):
    if len(argv) != 2:
        print(__doc__)
        return 2

    pasta = Path(argv[1])
    if not pasta.is_dir():
        print(f'Pasta não encontrada: {pasta}')
        return 1

    faltando = [n for n, *_ in
                [(f'CREG{i}.xlsx',) for i in (1, 2, 3, 4)] + [(p[0],) for p in PAGINAS]
                if not (pasta / n).is_file()]
    if faltando:
        print('Planilhas ausentes: ' + ', '.join(sorted(set(faltando))))
        return 1

    saida = Path(__file__).resolve().parent

    lidas_a, perdidas_a = ler_acervo(pasta)
    lidas_j, perdidas_j = ler_julgados(pasta)

    # A unificação de grafia roda sobre as duas listas juntas: o mesmo assunto
    # tem de sair escrito igual no acervo e nos julgados, senão um relatório que
    # cruze as duas tabelas volta a ter duas categorias.
    unificar(lidas_a + lidas_j, 'assunto')
    unificar(lidas_j, 'voto')
    unificar(lidas_j, 'status')

    acervo, repet_a = deduplicar(
        lidas_a, ('num_processo', 'data_distribuicao', 'unidade'))
    julgados, repet_j = deduplicar(
        lidas_j, ('num_processo', 'data_sessao'))

    (saida / 'acervo_creg.sql').write_text(gerar_sql(
        'acervo_creg',
        ['num_processo', 'unidade', 'data_distribuicao', 'assunto', 'recurso', 'origem'],
        acervo, 'acervo_creg_distribuicao_unica',
        f'Acervo do Conselho Regulador — {len(acervo)} distribuições (CREG1..4).',
    ), encoding='utf-8')

    (saida / 'julgados_creg.sql').write_text(gerar_sql(
        'julgados_creg',
        ['num_processo', 'data_sessao', 'pauta', 'voto', 'status', 'unidade',
         'data_distribuicao', 'assunto', 'recurso', 'defesa', 'data_dist_cj',
         'relator_cj', 'voto_cj'],
        julgados, 'julgados_creg_sessao_unica',
        f'Julgados do Conselho Regulador — {len(julgados)} sessões '
        f'({len({l["data_sessao"] for l in julgados})} datas).',
    ), encoding='utf-8')

    print(f'acervo_creg.sql   {len(acervo):5d} linhas  ({repet_a} cópias descartadas)')
    print(f'julgados_creg.sql {len(julgados):5d} linhas  ({repet_j} cópias descartadas)')

    # Linha que a planilha tinha e o banco não vai receber. Sair em silêncio
    # deixaria uma carga incompleta parecendo completa.
    perdidas = perdidas_a + perdidas_j
    if perdidas:
        print(f'\nATENÇÃO: {len(perdidas)} linha(s) com dado inaproveitável, '
              'fora da carga:')
        for d in perdidas:
            print(f'  {d}')

    print(f'\nRode acervo_creg.sql antes de julgados_creg.sql. Pasta: {saida}')
    return 0


if __name__ == '__main__':
    sys.exit(main(sys.argv))
