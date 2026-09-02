#!/usr/bin/env python3
"""Testes da Câmara de Julgamento contra um Postgres de verdade.

    python tests/test_cj.py ["C:/caminho/Câmara de Julgamento - REG.xlsx"]

Sobe um container postgres descartável (é o mesmo motor do Supabase), aplica
schema.sql, importa a planilha e confere que o banco reproduz
as regras que hoje só existem nas fórmulas. Sem a planilha, os testes que
dependem dela são pulados e o resto roda igual.

Requisitos: docker e psycopg2.
"""

import json
import re
import subprocess
import sys
from datetime import date, datetime
from pathlib import Path

import psycopg2
import psycopg2.extras

RAIZ = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(Path(__file__).resolve().parent))
import banco            # noqa: E402
from banco import uma  # noqa: E402
sys.path.insert(0, str(RAIZ / 'dados'))  # importar_planilha, carregado sob demanda

PLANILHA_PADRAO = Path.home() / 'Downloads' / 'Câmara de Julgamento - REG.xlsx'
PG = banco.Postgres('sorteio_sei_test', 55433)
MIGRACAO = RAIZ / 'supabase' / 'migrations' / \
    '20260823165725_corrigir_integridade_creg_e_privilegios.sql'
# Converte relator de nome para cadeira. Rodada em preparar_banco DEPOIS da
# planilha, porque e o dado importado que ela tem de alcancar.
MIGRACAO_CADEIRAS = RAIZ / 'supabase' / 'migrations' / '20260824180000_cadeiras_cj.sql'
MIGRACAO_HARDENING = RAIZ / 'supabase' / 'migrations' / \
    '20260827104200_restringir_cadeiras_e_ping.sql'
MIGRACAO_INTERESSADO = RAIZ / 'supabase' / 'migrations' / \
    '20260827160000_interessado_creg.sql'
# Derruba a tabela do sorteio antigo do Conselho, mas só depois de conferir que
# o que ela guardava está em acervo_creg. Rodada à parte num teste, para provar
# que a conferência é de verdade e não decoração.
MIGRACAO_REMOCAO = RAIZ / 'supabase' / 'migrations' / \
    '20260902120000_remover_processos_sorteados.sql'

testes = []


def teste(fn):
    testes.append(fn)
    return fn


def rodar_arquivo(caminho):
    PG.rodar_arquivo(caminho)


def comando_da_migracao(tabela):
    """O UPDATE de nome para cadeira, extraído do arquivo da migração.

    Serve para que o teste exercite o comando que roda de verdade em produção,
    em vez de uma cópia que continuaria verde depois de a migração quebrar.
    """
    fonte = MIGRACAO_CADEIRAS.read_text(encoding='utf-8')
    return re.search(rf'update public\.{tabela}\b.*?;', fonte, re.S).group(0)


# ── Planilha ─────────────────────────────────────────────────────────────────

def dias_excel(v):
    """"Dias DT" é um número de dias que a planilha exibe formatado como data.

    Volta ao número desfazendo o serial do Excel — inclusive o 29/02/1900 que
    ele inventou e as diferenças negativas, quando a sessão vem antes da
    distribuição (a planilha mostra uma data de 1899 nesses casos).
    """
    if not isinstance(v, datetime):
        return v
    # Importado aqui, e não no topo: openpyxl só é necessário para os testes que
    # leem a planilha. No topo, a ausência dele derrubava a suíte inteira — o
    # contrário do que o cabeçalho promete ("sem a planilha o resto roda igual").
    from openpyxl.utils.datetime import to_excel
    return int(to_excel(v))


class Planilha:
    """Valores JÁ CALCULADOS pelas fórmulas — é contra eles que o banco é medido."""

    def __init__(self, caminho):
        # Ambos só existem por causa da planilha, e importar_planilha carrega
        # openpyxl no topo. Importar aqui é o que mantém a suíte de pé numa
        # máquina sem openpyxl instalado, que é o caso de quem não tem a planilha.
        import openpyxl
        import importar_planilha as imp
        wb = openpyxl.load_workbook(caminho, data_only=True)
        self.acervo, _ = imp.deduplicar(
            imp.ler_acervo(wb['Acervo']), ('num_processo', 'data_distribuicao', 'relator'))
        self.julgados, _ = imp.deduplicar(
            imp.ler_julgados(wb['Julgados']), ('num_processo', 'data_sessao'))

        ws = wb['Julgados']
        self.calculados = {}
        for r in range(2, ws.max_row + 1):
            num = imp.processo(ws.cell(r, 1).value)
            sessao = imp.dia(ws.cell(r, 4).value)
            if num is None or sessao is None:
                continue
            self.calculados.setdefault(
                (num, sessao),
                (dias_excel(ws.cell(r, 8).value), imp.texto(ws.cell(r, 10).value)))
        wb.close()


PLANILHA = None


def exige_planilha(fn):
    fn.precisa_planilha = True
    return fn


# ── Testes: estrutura ────────────────────────────────────────────────────────

@teste
def tabelas_criadas(cur):
    """acervo_cj e julgados_cj existem, com as colunas e tipos esperados."""
    esperado_acervo = {
        'id': 'bigint', 'num_processo': 'text',
        'relator': 'text', 'data_distribuicao': 'date', 'defesa': 'boolean',
        'assunto': 'text', 'recurso': 'text', 'ordem': 'integer',
        'sorteado_em': 'timestamp with time zone', 'origem': 'text',
        'criado_em': 'timestamp with time zone',
    }
    esperado_julgados = {
        'id': 'bigint', 'acervo_id': 'bigint', 'num_processo': 'text',
        'data_sessao': 'date', 'pauta': 'integer',
        'voto': 'text', 'status': 'text', 'defesa': 'boolean', 'relator': 'text',
        'data_distribuicao': 'date', 'dias_dt': 'integer', 'periodo_dt': 'text',
        'criado_em': 'timestamp with time zone',
        'atualizado_em': 'timestamp with time zone', 'atualizado_por': 'text',
    }
    for tabela, esperado in [('acervo_cj', esperado_acervo), ('julgados_cj', esperado_julgados)]:
        cur.execute("""select column_name, data_type from information_schema.columns
                        where table_schema = 'public' and table_name = %s""", (tabela,))
        assert dict(cur.fetchall()) == esperado, tabela


@teste
def schema_provisiona_marco_da_serie(cur):
    assert uma(cur, """select count(*), min(data_sessao) from pautas_cj
                        where url = 'marco:inicio-da-serie'""") == (1, date(2026, 6, 18))


@teste
def restricoes_e_indices(cur):
    """Chaves, unicidade e FK no lugar — é o que impede duplicação e órfão."""
    cur.execute("""select conname, contype from pg_constraint
                    where conrelid in ('public.acervo_cj'::regclass,
                                       'public.julgados_cj'::regclass)""")
    nomes = dict(cur.fetchall())
    assert nomes.get('acervo_cj_distribuicao_unica') == 'u'
    assert nomes.get('julgados_cj_sessao_unica') == 'u'
    assert 'f' in nomes.values(), 'julgados_cj precisa da FK para acervo_cj'

    assert uma(cur, """select confdeltype from pg_constraint
                        where conrelid = 'public.julgados_cj'::regclass
                          and contype = 'f'""") == 'a', 'FK não pode apagar em cascata'

    assert uma(cur, """select count(*) from pg_indexes
                        where tablename = 'julgados_cj' and indexname = 'idx_julgados_cj_acervo'""") == 1


@teste
def rls_da_cada_tabela_o_minimo(cur):
    """Sorteio só insere; julgados só é lido; pautas nem isso.

    A leitura de julgados_cj é a única porta aberta para o navegador, porque a
    página de registro precisa listar os pendentes. Gravar voto e status passa
    pela função registrar_votos, nunca por UPDATE direto — por isso não pode
    existir política de UPDATE nem de DELETE em lugar nenhum.
    """
    esperado = {
        'acervo_cj': {'INSERT'},
        'julgados_cj': {'SELECT'},
        'pautas_cj': set(),
        'cadeiras_cj': set(),
        # O Conselho Regulador tem o mesmo desenho, e por isso os mesmos limites.
        'acervo_creg': {'INSERT'},
        'julgados_creg': {'SELECT'},
        'pautas_creg': set(),
    }
    for tabela, comandos in esperado.items():
        assert uma(cur, 'select relrowsecurity from pg_class where oid = %s::regclass',
                   (f'public.{tabela}',)) is True, tabela
        cur.execute('select cmd, roles::text from pg_policies where tablename = %s', (tabela,))
        politicas = cur.fetchall()
        assert {cmd for cmd, _ in politicas} == comandos, tabela
        assert all('authenticated' in roles for _, roles in politicas), tabela


@teste
def navegador_nao_atualiza_nem_apaga_julgados(cur):
    """Nenhuma política de UPDATE ou DELETE em nenhuma tabela do sistema."""
    cur.execute("""select tablename, cmd from pg_policies
                    where schemaname = 'public' and cmd in ('UPDATE', 'DELETE', 'ALL')""")
    assert cur.fetchall() == []


@teste
def privilegios_sql_repetem_o_minimo_da_rls(cur):
    """Grants padrão do Supabase não podem ampliar o que cada política permite."""
    esperado = {
        ('authenticated', 'acervo_cj'): {'INSERT'},
        ('authenticated', 'julgados_cj'): {'SELECT'},
        ('authenticated', 'acervo_creg'): {'INSERT'},
        ('authenticated', 'julgados_creg'): {'SELECT'},
    }
    cur.execute("""select grantee, table_name, privilege_type
                     from information_schema.role_table_grants
                    where table_schema = 'public'
                      and table_name in ('acervo_cj', 'julgados_cj', 'pautas_cj',
                                         'cadeiras_cj', 'acervo_creg',
                                         'julgados_creg', 'pautas_creg')
                      and grantee in ('anon', 'authenticated')""")
    obtido = {}
    for papel, tabela, privilegio in cur.fetchall():
        obtido.setdefault((papel, tabela), set()).add(privilegio)
    assert obtido == esperado

    cur.execute("""select grantee, object_name, privilege_type
                     from information_schema.role_usage_grants
                    where object_schema = 'public' and object_type = 'SEQUENCE'
                      and grantee in ('anon', 'authenticated')""")
    assert set(cur.fetchall()) == {
        ('authenticated', 'acervo_cj_id_seq', 'USAGE'),
        ('authenticated', 'acervo_creg_id_seq', 'USAGE'),
    }


@teste
def gatilho_roda_com_privilegio_proprio(cur):
    """SECURITY DEFINER com search_path fixo: lê o acervo sem abrir a tabela."""
    cur.execute("""select prosecdef, proconfig,
                          has_function_privilege('anon', oid, 'execute'),
                          has_function_privilege('authenticated', oid, 'execute')
                     from pg_proc
                    where proname = 'julgados_cj_derivar_do_acervo'""")
    secdef, config, pode_anonimo, pode_autenticado = cur.fetchone()
    assert secdef is True
    assert config and any(c.startswith('search_path=') for c in config)
    assert pode_anonimo is False and pode_autenticado is False


@teste
def ping_tem_search_path_fixo_e_privilegio_minimo(cur):
    cur.execute("""select proconfig,
                          has_function_privilege('anon', oid, 'execute'),
                          has_function_privilege('authenticated', oid, 'execute'),
                          has_function_privilege('service_role', oid, 'execute')
                     from pg_proc
                    where oid = 'public.ping()'::regprocedure""")
    config, pode_anonimo, pode_autenticado, pode_servico = cur.fetchone()
    assert config and any(c.startswith('search_path=') for c in config)
    assert pode_anonimo is True and pode_autenticado is True
    assert pode_servico is False

    assert uma(cur, """select count(*) from pg_proc p
                        join pg_namespace n on n.oid = p.pronamespace
                          where n.nspname = 'public'
                            and p.proname = 'painel_cj_nao_julgados'""") == 0


@teste
def tabela_antiga_foi_removida(cur):
    """processos_sorteados saiu do schema, e com ela tudo que pendurava nela.

    Era a tabela do sorteio antigo do Conselho: sem acervo, sem julgados e sem
    vínculo com nada, provisória até o CREG ganhar o desenho da Câmara. O que
    ela guardava vive em acervo_creg desde a migração 20260828090000, e a
    20260902120000 a derrubou.

    Não basta a tabela ter sumido: sequência, política de RLS e privilégio
    sobrevivem a um drop malfeito e continuariam aparecendo em auditoria.
    """
    assert uma(cur, "select to_regclass('public.processos_sorteados')") is None
    assert uma(cur, "select to_regclass('public.processos_sorteados_id_seq')") is None

    assert uma(cur, """select count(*) from pg_policies
                        where schemaname = 'public'
                          and tablename = 'processos_sorteados'""") == 0
    assert uma(cur, """select count(*) from information_schema.role_table_grants
                        where table_schema = 'public'
                          and table_name = 'processos_sorteados'""") == 0

    # O índice e a restrição de 15 dígitos caíram junto com a tabela.
    assert uma(cur, """select count(*) from pg_indexes
                        where schemaname = 'public'
                          and indexname like '%processos_sorteados%'""") == 0
    assert uma(cur, """select count(*) from pg_constraint
                        where conname like '%processos_sorteados%'""") == 0


# ── Testes: importação da planilha ───────────────────────────────────────────

@teste
@exige_planilha
def importacao_do_acervo(cur):
    """Toda distribuição da aba Acervo está no banco, com os valores certos."""
    assert uma(cur, "select count(*) from acervo_cj where origem = 'planilha'") == len(PLANILHA.acervo)

    cadeira = como_cadeira(cur)
    esperado = {(l['num_processo'], l['data_distribuicao'],
                 cadeira(l['relator'], l['data_distribuicao'])): l['defesa']
                for l in PLANILHA.acervo}
    cur.execute("""select num_processo, data_distribuicao, relator, defesa
                     from acervo_cj where origem = 'planilha'""")
    assert {(n, d, r): f for n, d, r, f in cur.fetchall()} == esperado


@teste
@exige_planilha
def importacao_dos_julgados(cur):
    """Toda sessão da aba Julgados está no banco, com os valores certos."""
    assert uma(cur, 'select count(*) from julgados_cj') == len(PLANILHA.julgados)

    cur.execute("""select num_processo, data_sessao, pauta, voto, status,
                          defesa, relator, data_distribuicao from julgados_cj""")
    banco = {(n, s): (p, v, st, d, r, dd) for n, s, p, v, st, d, r, dd in cur.fetchall()}
    cadeira = como_cadeira(cur)
    for l in PLANILHA.julgados:
        chave = (l['num_processo'], l['data_sessao'])
        # Em julgados_cj a migração decide pela data da SESSÃO, não pela da
        # distribuição — e o histórico tem linha com distribuição posterior à
        # sessão, onde os dois critérios divergem.
        assert banco[chave] == (l['pauta'], l['voto'], l['status'], l['defesa'],
                                cadeira(l['relator'], l['data_sessao']),
                                l['data_distribuicao']), chave


@teste
@exige_planilha
def importar_de_novo_nao_duplica(cur):
    """Reimportar as duas abas é no-op: a chave natural segura."""
    antes = (uma(cur, 'select count(*) from acervo_cj'),
             uma(cur, 'select count(*) from julgados_cj'))
    rodar_arquivo(RAIZ / 'dados' / 'acervo_cj.sql')
    rodar_arquivo(RAIZ / 'dados' / 'julgados_cj.sql')
    assert (uma(cur, 'select count(*) from acervo_cj'),
            uma(cur, 'select count(*) from julgados_cj')) == antes


@teste
@exige_planilha
def sem_duplicidade_de_chave_natural(cur):
    """Nenhuma distribuição repetida no acervo, nenhum processo julgado 2x na sessão."""
    assert uma(cur, """select count(*) from (
                         select 1 from acervo_cj
                          group by num_processo, data_distribuicao, relator
                         having count(*) > 1) x""") == 0
    assert uma(cur, """select count(*) from (
                         select 1 from julgados_cj
                          group by num_processo, data_sessao
                         having count(*) > 1) x""") == 0


# ── Testes: relacionamento e regras derivadas ────────────────────────────────

@teste
@exige_planilha
def todo_julgado_aponta_para_o_acervo(cur):
    """A ligação existe e é coerente: o processo do julgado é o do acervo."""
    assert uma(cur, 'select count(*) from julgados_cj where acervo_id is null') == 0
    assert uma(cur, """select count(*) from julgados_cj j
                       join acervo_cj a on a.id = j.acervo_id
                      where a.num_processo <> j.num_processo""") == 0


@teste
@exige_planilha
def dias_dt_bate_com_a_formula(cur):
    """dias_dt (coluna gerada) == "Dias DT" (=-I+D) para todas as linhas."""
    cur.execute('select num_processo, data_sessao, dias_dt from julgados_cj')
    divergentes = [(n, s) for n, s, d in cur.fetchall()
                   if PLANILHA.calculados[(n, s)][0] != d]
    assert not divergentes, divergentes[:5]


@teste
@exige_planilha
def periodo_dt_bate_com_a_formula(cur):
    """periodo_dt (coluna gerada) == "Per DT" (IF aninhado por trimestre)."""
    cur.execute('select num_processo, data_sessao, periodo_dt from julgados_cj')
    divergentes = [(n, s, p, PLANILHA.calculados[(n, s)][1]) for n, s, p in cur.fetchall()
                   if PLANILHA.calculados[(n, s)][1] != p]
    assert not divergentes, divergentes[:5]


@teste
def periodo_dt_cobre_o_que_a_planilha_nao_cobria(cur):
    """A planilha parava em 2026 e não tratava trimestre antes de 2023 direito."""
    cur.execute("""insert into acervo_cj (num_processo, relator, data_distribuicao, origem)
                   values ('900000000000001', 'CJ9', date '2021-01-05', 'sorteio')""")
    cur.execute("""insert into julgados_cj (num_processo, data_sessao) values
                   ('900000000000001', date '2022-02-10'),
                   ('900000000000001', date '2027-11-30'),
                   ('900000000000001', date '2030-07-01')
                   returning periodo_dt""")
    assert [r[0] for r in cur.fetchall()] == ['<22', '4T27', '3T30']
    cur.connection.rollback()


# ── Testes: preenchimento automático ao registrar um julgamento ──────────────

@teste
def preenche_a_partir_do_acervo(cur):
    """O fluxo do dia a dia: informa processo e sessão, o banco busca o resto."""
    cur.execute("""insert into acervo_cj
                   (num_processo, relator, data_distribuicao, defesa, origem)
                   values ('900000000000010', 'CJ3', date '2026-03-02', true, 'sorteio')""")
    cur.execute("""insert into julgados_cj (num_processo, data_sessao, pauta, voto, status)
                   values ('900000000000010', date '2026-06-10', 7, 'Manter', 'Julgado')
                   returning relator, defesa, data_distribuicao,
                             dias_dt, periodo_dt, acervo_id is not null""")
    assert cur.fetchone() == ('CJ3', True, date(2026, 3, 2), 100, '2T26', True)
    cur.connection.rollback()


@teste
def usa_a_distribuicao_vigente_na_data_da_sessao(cur):
    """Processo redistribuído: vale o relator que tinha o processo na sessão."""
    cur.execute("""insert into acervo_cj (num_processo, relator, data_distribuicao, defesa, origem)
                   values ('900000000000011', 'CJ1', date '2026-01-10', false, 'sorteio'),
                          ('900000000000011', 'CJ2', date '2026-04-20', true,  'sorteio'),
                          ('900000000000011', 'CJ4', date '2026-09-30', false, 'sorteio')""")
    cur.execute("""insert into julgados_cj (num_processo, data_sessao)
                   values ('900000000000011', date '2026-05-05')
                   returning relator, defesa, data_distribuicao""")
    assert cur.fetchone() == ('CJ2', True, date(2026, 4, 20))
    cur.connection.rollback()


@teste
def distribuicao_informada_manda_no_vinculo(cur):
    """Ao importar histórico, a data informada escolhe a distribuição exata."""
    cur.execute("""insert into acervo_cj (num_processo, relator, data_distribuicao, origem)
                   values ('900000000000012', 'CJ1', date '2026-01-10', 'sorteio'),
                          ('900000000000012', 'CJ2', date '2026-04-20', 'sorteio')""")
    cur.execute("""insert into julgados_cj (num_processo, data_sessao, data_distribuicao)
                   values ('900000000000012', date '2026-05-05', date '2026-01-10')
                   returning relator, (select relator from acervo_cj a where a.id = acervo_id)""")
    assert cur.fetchone() == ('CJ1', 'CJ1')
    cur.connection.rollback()


@teste
def distribuicao_informada_sem_correspondencia_nao_inventa_vinculo(cur):
    """Data manual sem distribuição igual deve ficar órfã para revisão."""
    cur.execute("""insert into acervo_cj (num_processo, relator, data_distribuicao, origem)
                   values ('900000000000034', 'CJ5', date '2026-08-15', 'sorteio')""")
    cur.execute("""insert into julgados_cj
                   (num_processo, data_sessao, data_distribuicao, relator, defesa)
                   values ('900000000000034', date '2026-08-20', date '2026-08-01',
                           'Conselheira Manual', false)
                   returning acervo_id, data_distribuicao, relator, defesa""")
    assert cur.fetchone() == (None, date(2026, 8, 1), 'Conselheira Manual', False)
    cur.connection.rollback()


@teste
def distribuicao_posterior_a_sessao_ainda_e_encontrada(cur):
    """Caso da planilha: única distribuição é posterior à sessão (INDEX/MATCH acha)."""
    cur.execute("""insert into acervo_cj (num_processo, relator, data_distribuicao, origem)
                   values ('900000000000013', 'CJ5', date '2026-08-07', 'sorteio')""")
    cur.execute("""insert into julgados_cj (num_processo, data_sessao)
                   values ('900000000000013', date '2026-01-25')
                   returning relator, data_distribuicao, dias_dt""")
    assert cur.fetchone() == ('CJ5', date(2026, 8, 7), -194)
    cur.connection.rollback()


@teste
def valor_informado_vence_o_derivado(cur):
    """As 1.122 linhas com Defesa digitada à mão não podem ser sobrescritas."""
    cur.execute("""insert into acervo_cj (num_processo, relator, data_distribuicao, defesa, origem)
                   values ('900000000000014', 'CJ1', date '2026-02-01', false, 'sorteio')""")
    cur.execute("""insert into julgados_cj (num_processo, data_sessao, defesa, relator)
                   values ('900000000000014', date '2026-03-01', true, 'Conselheiro Fulano')
                   returning defesa, relator""")
    assert cur.fetchone() == (True, 'Conselheiro Fulano')
    cur.connection.rollback()


@teste
def limpar_campo_faz_o_banco_rederivar(cur):
    """Gravar null num campo derivado é como pedir a fórmula de volta."""
    cur.execute("""insert into acervo_cj (num_processo, relator, data_distribuicao, origem)
                   values ('900000000000015', 'CJ2', date '2026-02-01', 'sorteio')""")
    cur.execute("""insert into julgados_cj (num_processo, data_sessao, relator)
                   values ('900000000000015', date '2026-03-01', 'Errado')""")
    cur.execute("""update julgados_cj set relator = null
                    where num_processo = '900000000000015' returning relator""")
    assert cur.fetchone()[0] == 'CJ2'
    cur.connection.rollback()


@teste
def processo_fora_do_acervo_nao_quebra(cur):
    """Equivale ao "Não encontrado" da planilha: registra sem vínculo."""
    cur.execute("""insert into julgados_cj (num_processo, data_sessao, pauta)
                   values ('900000000000099', date '2026-06-10', 3)
                   returning acervo_id, relator, defesa, data_distribuicao, dias_dt, periodo_dt""")
    assert cur.fetchone() == (None, None, None, None, None, '2T26')
    cur.connection.rollback()


@teste
def rederivar_liga_o_julgado_ao_acervo_que_chegou_depois(cur):
    """Sortear um processo já julgado não conserta o julgado sozinho.

    O gatilho só dispara em julgados_cj, e inserir no acervo não toca nela — o
    julgado continua órfão por mais que a distribuição já exista. O
    rederivar_cj.sql é o empurrão que refaz o vínculo, sem encostar em voto e
    status e sem mexer em quem já estava vinculado.
    """
    cur.execute("""insert into acervo_cj
                   (num_processo, relator, data_distribuicao, defesa, origem)
                   values ('900000000000031', 'CJ1', date '2026-07-01', true, 'sorteio')""")
    cur.execute("""insert into julgados_cj (num_processo, data_sessao, voto, status) values
                     ('900000000000031', date '2026-08-06', 'Manter', 'Julgado'),
                     ('900000000000032', date '2026-08-13', 'Anular', 'Julgado')""")

    # O sorteio do 032 só acontece agora, depois de ele já ter ido à sessão.
    cur.execute("""insert into acervo_cj
                   (num_processo, relator, data_distribuicao, defesa, origem)
                   values ('900000000000032', 'CJ4', date '2026-07-15', true, 'sorteio')""")
    assert uma(cur, """select acervo_id from julgados_cj
                        where num_processo = '900000000000032'""") is None, \
        'inserir no acervo não deveria consertar o julgado sozinho'

    cur.execute((RAIZ / 'sql' / 'rederivar_cj.sql').read_text(encoding='utf-8'))
    relatorio = {num: (resultado, relator) for num, _, resultado, relator, _, _ in cur.fetchall()}

    assert relatorio.get('900000000000032') == ('vinculado agora', 'CJ4')
    assert '900000000000031' not in relatorio, 'julgado já vinculado não pode ser tocado'

    # Voto e status são da sessão, não do acervo: a rederivação não os alcança.
    assert uma(cur, """select acervo_id is not null, relator, defesa, dias_dt, voto, status
                         from julgados_cj where num_processo = '900000000000032'""") == \
        (True, 'CJ4', True, 29, 'Anular', 'Julgado')

    cur.connection.rollback()


@teste
def rederivar_vincula_orfao_sem_apagar_campos_manuais(cur):
    """O vínculo faltante volta sem trocar os valores já revisados pela secretaria."""
    cur.execute("""insert into julgados_cj
                   (num_processo, data_sessao, relator, defesa, data_distribuicao)
                   values ('900000000000033', date '2026-08-20',
                           'Conselheira Manual', false, date '2026-08-01')""")
    cur.execute("""insert into acervo_cj
                   (num_processo, relator, data_distribuicao, defesa, origem)
                   values ('900000000000033', 'CJ2', date '2026-08-01', true, 'sorteio'),
                          ('900000000000033', 'CJ5', date '2026-08-15', true, 'sorteio')""")

    cur.execute((RAIZ / 'sql' / 'rederivar_cj.sql').read_text(encoding='utf-8'))

    assert uma(cur, """select j.acervo_id is not null, j.relator, j.defesa,
                              j.data_distribuicao, j.dias_dt, a.data_distribuicao
                         from julgados_cj j join acervo_cj a on a.id = j.acervo_id
                        where j.num_processo = '900000000000033'""") == \
        (True, 'Conselheira Manual', False, date(2026, 8, 1), 19, date(2026, 8, 1))
    cur.connection.rollback()


@teste
def campos_opcionais_aceitam_vazio(cur):
    """Voto, status e pauta podem faltar — a planilha tem linhas assim."""
    cur.execute("""insert into acervo_cj (num_processo, relator, data_distribuicao, origem)
                   values ('900000000000016', 'CJ1', date '2026-02-01', 'sorteio')""")
    cur.execute("""insert into julgados_cj (num_processo, data_sessao)
                   values ('900000000000016', date '2026-03-01')
                   returning voto, status, pauta""")
    assert cur.fetchone() == (None, None, None)
    cur.connection.rollback()


@teste
def sessao_repetida_e_barrada(cur):
    """Mesmo processo, mesma sessão, duas vezes: o banco recusa."""
    cur.execute("""insert into acervo_cj (num_processo, relator, data_distribuicao, origem)
                   values ('900000000000017', 'CJ1', date '2026-02-01', 'sorteio')""")
    cur.execute("""insert into julgados_cj (num_processo, data_sessao)
                   values ('900000000000017', date '2026-03-01')""")
    try:
        cur.execute("""insert into julgados_cj (num_processo, data_sessao)
                       values ('900000000000017', date '2026-03-01')""")
    except psycopg2.errors.UniqueViolation:
        return
    finally:
        cur.connection.rollback()
    raise AssertionError('julgados_cj aceitou a mesma sessão duas vezes')


@teste
def sorteio_repetido_e_barrado(cur):
    """Mesmo processo, mesmo dia, mesma cadeira: o acervo recusa."""
    cur.execute("""insert into acervo_cj (num_processo, relator, data_distribuicao, origem)
                   values ('900000000000018', 'CJ1', date '2026-02-01', 'sorteio')""")
    try:
        cur.execute("""insert into acervo_cj (num_processo, relator, data_distribuicao, origem)
                       values ('900000000000018', 'CJ1', date '2026-02-01', 'sorteio')""")
    except psycopg2.errors.UniqueViolation:
        return
    finally:
        cur.connection.rollback()
    raise AssertionError('acervo_cj aceitou a mesma distribuição duas vezes')


@teste
def creg_recusa_a_mesma_distribuicao_duas_vezes(cur):
    """Mesmo processo, dia e unidade do CREG não podem virar duas linhas.

    A regra saiu da tabela antiga junto com ela: quem a carrega hoje é a
    restrição acervo_creg_distribuicao_unica.
    """
    valores = ("202600029000020", "CREG1", "2026-08-22")
    cur.execute("""insert into acervo_creg
                   (num_processo, unidade, data_distribuicao, assunto, recurso, ordem)
                   values (%s, %s, %s, 'Requerimento', 'Não se aplica', 1)""",
                valores)
    try:
        cur.execute("""insert into acervo_creg
                       (num_processo, unidade, data_distribuicao, assunto, recurso, ordem)
                       values (%s, %s, %s, 'Requerimento', 'Não se aplica', 2)""",
                    valores)
    except psycopg2.errors.UniqueViolation:
        return
    finally:
        cur.connection.rollback()
    raise AssertionError('acervo_creg aceitou a mesma distribuição duas vezes')


@teste
def creg_novo_exige_numero_de_processo_com_15_digitos(cur):
    """O acervo do Conselho recusa processo fora do padrão SEI.

    Na tabela antiga isso era restrição acrescentada depois, NOT VALID até o
    legado ser limpo. Em acervo_creg nasce no CREATE TABLE: não há legado a
    tolerar, então vale para toda linha desde a primeira.
    """
    try:
        cur.execute("""insert into acervo_creg
                       (num_processo, unidade, data_distribuicao, assunto, recurso)
                       values ('1234', 'CREG3', current_date,
                               'Requerimento', 'Não se aplica')""")
    except psycopg2.errors.CheckViolation:
        return
    finally:
        cur.connection.rollback()
    raise AssertionError('novo CREG com processo fora de 15 dígitos foi aceito')


@teste
def migracao_levou_o_sorteio_antigo_para_o_acervo(cur):
    """A cadeia inteira que aposentou a tabela antiga, medida no resultado.

    Os fixtures nascem em processos_sorteados (preparar_upgrade_da_migracao),
    com a data do sorteio de 27/08/2026. Daí em diante, três migrações em
    sequência: a 20260823165725 apaga os dois registros fora do padrão, a
    20260828090000 copia o que sobrou para acervo_creg e a 20260902120000
    derruba a tabela — só depois de conferir que nada ficou para trás.

    Se qualquer um dos três elos falhar, é aqui que aparece: o processo válido
    tem de estar no acervo, com os valores que tinha, e os dois inválidos não
    podem ter atravessado a limpeza.
    """
    assert uma(cur, """select unidade, data_distribuicao, assunto, recurso,
                              ordem, origem
                         from acervo_creg
                        where num_processo = '202600029000777'""") == (
        'CREG2', date(2026, 8, 20), 'Requerimento', 'Com recurso', 3, 'sorteio')

    assert uma(cur, """select count(*) from acervo_creg
                        where num_processo in ('123421', '1234')""") == 0


@teste
def remocao_recusa_apagar_o_que_nao_foi_copiado(cur):
    """A trava da migração de remoção é de verdade: sem cópia, ela não apaga.

    Toda a segurança de derrubar uma tabela com dado dentro está nessa
    conferência. Se ela estivesse escrita errada — condição invertida, join que
    sempre casa — a migração passaria verde apagando o que ninguém teria como
    recuperar, e nenhum outro teste perceberia: no banco da bateria a cópia
    sempre deu certo.

    Então aqui a tabela volta com uma linha que não está em acervo_creg, e a
    migração roda de novo. Tem de parar, e a linha tem de continuar lá.
    """
    cur.execute("""create table public.processos_sorteados (
                     id                bigint generated always as identity primary key,
                     modo              text not null,
                     data_hora         timestamptz not null,
                     ordem             int,
                     num_processo      text not null,
                     assunto           text,
                     data_distribuicao date not null,
                     recurso           text,
                     unidade           text not null)""")
    cur.execute("""insert into public.processos_sorteados
                   (modo, data_hora, ordem, num_processo, assunto,
                    data_distribuicao, recurso, unidade)
                   values ('CREG', now(), 1, '202600029000888', 'Requerimento',
                           date '2026-08-20', 'Com recurso', 'CREG4')""")
    cur.connection.commit()

    try:
        try:
            rodar_arquivo(MIGRACAO_REMOCAO)
        except psycopg2.errors.RaiseException as erro:
            assert 'acervo_creg' in str(erro), erro
        else:
            raise AssertionError('a migração apagou a tabela sem conferir a cópia')

        sobrou = uma(cur, """select count(*) from public.processos_sorteados
                              where num_processo = '202600029000888'""")
        assert sobrou == 1, 'a linha sem cópia devia ter sobrevivido'
    finally:
        # A transação de cur segura um lock sobre a tabela; o drop vem de outra
        # conexão e ficaria esperando por ela para sempre. Fechar antes não é
        # zelo, é o que impede a bateria de travar aqui.
        cur.connection.rollback()
        PG.executar('drop table if exists public.processos_sorteados')

    assert uma(cur, "select to_regclass('public.processos_sorteados')") is None


@teste
def origem_so_aceita_valores_conhecidos(cur):
    try:
        cur.execute("""insert into acervo_cj (num_processo, relator, data_distribuicao, origem)
                       values ('900000000000019', 'CJ1', date '2026-02-01', 'chute')""")
    except psycopg2.errors.CheckViolation:
        return
    finally:
        cur.connection.rollback()
    raise AssertionError('origem aceitou valor fora da lista')


# ── Teste: as regras do banco == as fórmulas da planilha ─────────────────────

@teste
@exige_planilha
def regras_do_banco_reproduzem_as_formulas(cur):
    """O teste que importa: apaga tudo de julgados_cj e manda o banco derivar
    do zero as 3.144 linhas, informando só o que a sessão informa. Compara com o
    que as fórmulas calcularam.

    Divergência esperada e desejada: processo redistribuído. As fórmulas se
    contradizem — Relator/Defesa vêm do INDEX/MATCH (primeira distribuição) e
    Data DIST vem do AGGREGATE/MAIOR (última distribuição). O banco usa a mesma
    origem para os três campos: a distribuição vigente na data da sessão. Toda
    divergência abaixo tem que ser de processo com mais de uma distribuição.
    """
    cadeira = como_cadeira(cur)
    cur.execute('delete from julgados_cj')
    psycopg2.extras.execute_values(cur, """
        insert into julgados_cj (num_processo, data_sessao, pauta, voto, status)
        values %s""", [(l['num_processo'], l['data_sessao'],
                        l['pauta'], l['voto'], l['status']) for l in PLANILHA.julgados])

    cur.execute("""select num_processo, data_sessao, relator, defesa, data_distribuicao,
                          dias_dt, periodo_dt, acervo_id from julgados_cj""")
    derivado = {(n, s): resto for n, s, *resto in cur.fetchall()}

    cur.execute("""select num_processo from acervo_cj
                    group by num_processo having count(*) > 1""")
    redistribuidos = {r[0] for r in cur.fetchall()}

    achados = {'relator': [], 'defesa': [], 'data_distribuicao': []}
    for l in PLANILHA.julgados:
        chave = (l['num_processo'], l['data_sessao'])
        rel, dfs, ddist, dias, per, acervo_id = derivado[chave]

        dias_planilha, per_planilha = PLANILHA.calculados[chave]
        assert acervo_id is not None, f'{chave} deixou de achar o processo no acervo'
        assert per == per_planilha, chave
        # dias_dt é sempre a subtração da fórmula; quando a distribuição
        # escolhida é a mesma da planilha, o número também tem que ser o mesmo.
        assert dias == (l['data_sessao'] - ddist).days, chave
        assert ddist != l['data_distribuicao'] or dias == dias_planilha, chave

        # O relator da planilha é o nome; o do banco, a cadeira em que a
        # migração o converteu. Comparar sem traduzir acusaria as 3.144 linhas.
        for campo, obtido in (('relator', rel), ('defesa', dfs), ('data_distribuicao', ddist)):
            esperado_planilha = cadeira(l[campo], ddist) if campo == 'relator' else l[campo]
            if esperado_planilha != obtido:
                achados[campo].append((chave, l[campo], obtido))

    # Segunda divergência esperada: a aba Julgados tem 1.122 linhas com Defesa
    # digitada à mão e algumas contradizem o acervo. Aqui elas foram apagadas de
    # propósito, então o banco devolve o que o acervo diz. Na importação de
    # verdade o valor digitado é preservado — é o que o teste
    # valor_informado_vence_o_derivado garante.
    defesa_no_acervo = {}
    for a in PLANILHA.acervo:
        defesa_no_acervo.setdefault(a['num_processo'], set()).add(a['defesa'])

    manuais = 0
    for campo, casos in achados.items():
        fora = []
        for (num, sessao), da_planilha, do_banco in casos:
            if num in redistribuidos:
                continue
            if campo == 'defesa' and da_planilha not in defesa_no_acervo.get(num, set()):
                manuais += 1
                continue
            fora.append(((num, sessao), da_planilha, do_banco))
        assert not fora, f'{campo} divergiu sem explicação: {fora[:5]}'

    # Coerência interna que a planilha não tinha: os três campos derivados de um
    # julgado saem sempre do mesmo registro do acervo.
    assert uma(cur, """select count(*) from julgados_cj j join acervo_cj a on a.id = j.acervo_id
                        where a.relator is distinct from j.relator
                           or a.defesa is distinct from j.defesa
                           or a.data_distribuicao is distinct from j.data_distribuicao""") == 0

    print(f'      divergências vs. fórmulas em {len(PLANILHA.julgados)} linhas: ' +
          ', '.join(f'{c}={len(v)}' for c, v in achados.items()) +
          f' (sendo {manuais} de Defesa digitada à mão, o resto por redistribuição)')
    cur.connection.rollback()


# ── Testes: registro de voto e status pela secretaria ────────────────────────

def julgado_pendente(cur, num='900000000000200', sessao='2026-07-02', pauta=23):
    """Um julgado como a sincronização cria: sem voto e sem status."""
    cur.execute("""insert into acervo_cj (num_processo, relator, data_distribuicao, origem)
                   values (%s, 'CJ1', date '2026-03-10', 'sorteio')
                   on conflict do nothing""", (num,))
    cur.execute("""insert into julgados_cj (num_processo, data_sessao, pauta)
                   values (%s, %s, %s) returning id, voto, status""", (num, sessao, pauta))
    return cur.fetchone()


def registrar(cur, itens, email='secretaria@goias.gov.br'):
    if email:
        cur.execute("select set_config('request.jwt.claims', %s, true)",
                    (json.dumps({'email': email, 'role': 'authenticated',
                                 'sub': '00000000-0000-0000-0000-000000000001'}),))
    return uma(cur, 'select public.registrar_votos(%s::jsonb)', (json.dumps(itens),))


@teste
def julgado_novo_nasce_sem_voto_e_sem_status(cur):
    """É o estado que a página de registro procura."""
    ident, voto, status = julgado_pendente(cur)
    assert (voto, status) == (None, None)
    cur.connection.rollback()


@teste
def registrar_votos_preenche_os_dois_campos(cur):
    ident, _, _ = julgado_pendente(cur)
    assert registrar(cur, [{'id': ident, 'voto': 'Manter', 'status': 'Julgado'}]) == 1

    cur.execute("""select voto, status, atualizado_por, atualizado_em is not null
                     from julgados_cj where id = %s""", (ident,))
    assert cur.fetchone() == ('Manter', 'Julgado', 'secretaria@goias.gov.br', True)
    cur.connection.rollback()


@teste
def registrar_votos_aceita_preenchimento_parcial(cur):
    """Processo retirado de pauta tem status e não tem voto."""
    ident, _, _ = julgado_pendente(cur)
    assert registrar(cur, [{'id': ident, 'voto': '', 'status': 'Retirado'}]) == 1

    cur.execute('select voto, status from julgados_cj where id = %s', (ident,))
    assert cur.fetchone() == (None, 'Retirado')
    cur.connection.rollback()


@teste
def registrar_votos_recusa_rotulo_desconhecido(cur):
    ident, _, _ = julgado_pendente(cur)
    for item in ({'id': ident, 'voto': 'Talvez', 'status': 'Julgado'},
                 {'id': ident, 'voto': 'Manter', 'status': 'Arquivado'},
                 {'id': '; drop table julgados_cj; --', 'voto': 'Manter'},
                 {'voto': 'Manter'}):
        try:
            registrar(cur, [item])
        except psycopg2.errors.RaiseException:
            cur.connection.rollback()
            continue
        raise AssertionError(f'aceitou {item}')

    assert uma(cur, "select to_regclass('public.julgados_cj') is not null") is True


@teste
def registrar_votos_nao_encosta_no_historico_da_planilha(cur):
    """Linha já julgada e vinda da planilha é intocável por esta porta."""
    ident, _, _ = julgado_pendente(cur, num='900000000000201')
    cur.execute("""update julgados_cj set voto = 'Manter', status = 'Julgado',
                          atualizado_em = null, atualizado_por = null
                    where id = %s""", (ident,))

    assert registrar(cur, [{'id': ident, 'voto': 'Anular', 'status': 'Retirado'}]) == 0
    cur.execute('select voto, status from julgados_cj where id = %s', (ident,))
    assert cur.fetchone() == ('Manter', 'Julgado')
    cur.connection.rollback()


@teste
def registrar_votos_nao_apaga_decisao_com_campo_em_branco(cur):
    """Branco quer dizer "ainda não decidi", nunca "apague o que está lá".

    A função escrevia voto E status incondicionalmente: a linha que tem um dos
    dois preenchido entra na fila de pendentes justamente por isso, e gravar a
    sessão levava o campo já decidido junto. Pela API, {"voto":"","status":""}
    zerava as duas colunas de qualquer linha que a tela já tivesse encostado.
    """
    ident, _, _ = julgado_pendente(cur, num='900000000000203')
    cur.execute("""update julgados_cj set voto = 'Manter', status = null,
                          atualizado_em = null, atualizado_por = null
                    where id = %s""", (ident,))

    assert registrar(cur, [{'id': ident, 'voto': '', 'status': 'Julgado'}]) == 1
    cur.execute('select voto, status from julgados_cj where id = %s', (ident,))
    assert cur.fetchone() == ('Manter', 'Julgado')

    # Agora a linha é editável por esta porta, e o branco continua sem apagar.
    registrar(cur, [{'id': ident, 'voto': '', 'status': ''}])
    cur.execute('select voto, status from julgados_cj where id = %s', (ident,))
    assert cur.fetchone() == ('Manter', 'Julgado')
    cur.connection.rollback()


@teste
def registrar_votos_permite_corrigir_o_proprio_registro(cur):
    """O que a página gravou, a página conserta."""
    ident, _, _ = julgado_pendente(cur, num='900000000000202')
    registrar(cur, [{'id': ident, 'voto': 'Manter', 'status': 'Julgado'}])
    assert registrar(cur, [{'id': ident, 'voto': 'Anular', 'status': 'Julgado'}]) == 1

    cur.execute('select voto, status from julgados_cj where id = %s', (ident,))
    assert cur.fetchone() == ('Anular', 'Julgado')
    cur.connection.rollback()


@teste
def registrar_votos_so_mexe_em_voto_e_status(cur):
    ident, _, _ = julgado_pendente(cur, num='900000000000203')
    cur.execute("""select num_processo, data_sessao, pauta, relator, defesa,
                          data_distribuicao, acervo_id
                     from julgados_cj where id = %s""", (ident,))
    antes = cur.fetchone()

    registrar(cur, [{'id': ident, 'voto': 'Vista', 'status': 'Vista'}])

    cur.execute("""select num_processo, data_sessao, pauta, relator, defesa,
                          data_distribuicao, acervo_id
                     from julgados_cj where id = %s""", (ident,))
    assert cur.fetchone() == antes
    cur.connection.rollback()


@teste
def registrar_votos_grava_varios_de_uma_vez(cur):
    a, _, _ = julgado_pendente(cur, num='900000000000204')
    b, _, _ = julgado_pendente(cur, num='900000000000205')
    assert registrar(cur, [{'id': a, 'voto': 'Manter', 'status': 'Julgado'},
                           {'id': b, 'voto': 'Anular', 'status': 'Julgado'}]) == 2
    cur.connection.rollback()


@teste
def registrar_votos_recusa_jwt_sem_usuario(cur):
    ident, _, _ = julgado_pendente(cur, num='900000000000206')
    cur.execute("select set_config('request.jwt.claims', %s, true)",
                (json.dumps({'role': 'authenticated', 'email': 'intruso@example.org'}),))
    try:
        uma(cur, 'select public.registrar_votos(%s::jsonb)',
            (json.dumps([{'id': ident, 'voto': 'Manter', 'status': 'Julgado'}]),))
    except psycopg2.Error as erro:
        assert erro.pgcode == '28000'
        return
    finally:
        cur.connection.rollback()
    raise AssertionError('registrar_votos aceitou JWT sem usuário')


@teste
def registrar_votos_e_a_unica_porta_de_escrita(cur):
    """SECURITY DEFINER, search_path fixo e execução só para autenticado."""
    cur.execute("""select prosecdef, proconfig,
                          has_function_privilege('authenticated', oid, 'execute'),
                          has_function_privilege('anon', oid, 'execute'),
                          has_function_privilege('service_role', oid, 'execute')
                     from pg_proc where proname = 'registrar_votos'""")
    secdef, config, pode_autenticado, pode_anonimo, pode_servico = cur.fetchone()
    assert secdef is True
    assert config and any(c.startswith('search_path=') for c in config)
    assert pode_autenticado is True and pode_anonimo is False and pode_servico is False

    cur.execute("select proconfig from pg_proc where proname = 'auth_email'")
    config_auth_email = cur.fetchone()[0]
    assert config_auth_email and any(c.startswith('search_path=') for c in config_auth_email)


# ── Testes: nada quebrou no que já existia ───────────────────────────────────

# ── Painel do acervo ─────────────────────────────────────────────────────────

def como_cadeira(cur):
    """Traduz nome de conselheiro para cadeira, como a migração faz.

    A suíte roda a migração das cadeiras depois de importar a planilha, para
    exercitá-la em CI. Os testes de fidelidade comparam o banco com a planilha,
    que traz o NOME — então o esperado precisa passar pelo mesmo de-para.

    E pela mesma REGRA: a conversão só alcança o que está dentro da vigência da
    cadeira. Uma linha de 2024 é de outra composição e fica pelo nome; traduzir
    sem olhar a data acusaria divergência justamente onde o banco acertou.
    """
    cur.execute('select conselheiro, cadeira, desde, ate from public.cadeiras_cj')
    periodos = cur.fetchall()

    def traduzir(relator, data):
        for conselheiro, cadeira, desde, ate in periodos:
            if conselheiro == relator and data >= desde and (ate is None or data <= ate):
                return cadeira
        return relator
    return traduzir


def autenticar(cur, email='secretaria@goias.gov.br'):
    cur.execute("select set_config('request.jwt.claims', %s, true)",
                (json.dumps({'email': email, 'role': 'authenticated',
                             'sub': '00000000-0000-0000-0000-000000000001'}),))


@teste
def resumo_do_acervo_exige_sessao(cur):
    """A função lê acervo_cj, que é fechada ao navegador.

    Sem a checagem, ela seria um endpoint /rest/v1/rpc que devolve o acervo
    inteiro agregado para quem tiver só a chave publicável.
    """
    cur.execute("select set_config('request.jwt.claims', '', true)")
    try:
        cur.execute('select * from public.resumo_acervo_cj()')
        raise AssertionError('respondeu sem sessão autenticada')
    except psycopg2.errors.InvalidAuthorizationSpecification:
        pass
    cur.connection.rollback()


@teste
def resumo_do_acervo_so_e_executavel_por_authenticated(cur):
    """EXECUTE é concedido a PUBLIC por padrão; sem o revoke, anon chamaria."""
    cur.execute("""select coalesce(string_agg(grantee, ',' order by grantee), '')
                     from information_schema.role_routine_grants
                    where routine_schema = 'public'
                      and routine_name = 'resumo_acervo_cj'
                      and grantee in ('anon', 'authenticated', 'service_role', 'PUBLIC')""")
    concedido = cur.fetchone()[0]
    assert concedido == 'authenticated', concedido


@teste
def resumo_do_acervo_conta_o_que_nao_foi_julgado(cur):
    """Processo julgado sai do painel; o que nunca foi a sessão fica."""
    autenticar(cur)
    cur.execute("delete from public.julgados_cj")
    cur.execute("delete from public.acervo_cj")
    cur.execute("""insert into public.acervo_cj
                     (num_processo, relator, data_distribuicao, defesa, origem) values
                   ('202600029000001', 'Fulano', current_date - 5,   true, 'sorteio'),
                   ('202600029000002', 'Fulano', current_date - 200, true, 'sorteio'),
                   ('202600029000003', 'Sicrano', current_date - 900, true, 'sorteio')""")
    # o 3 foi julgado: sai do painel
    cur.execute("""insert into public.julgados_cj (num_processo, data_sessao, pauta)
                   values ('202600029000003', current_date - 10, 1)""")

    cur.execute('select ordem, faixa, relator, processos from public.resumo_acervo_cj()')
    linhas = cur.fetchall()

    # a grade é sempre completa: 8 faixas x 2 relatores, com zero onde não há nada
    assert len(linhas) == 16, len(linhas)
    assert {r for _, _, r, _ in linhas} == {'Fulano', 'Sicrano'}
    assert sum(n for _, _, _, n in linhas) == 2, 'o julgado deveria ter saído'

    por_celula = {(f, r): n for _, f, r, n in linhas}
    assert por_celula[('Até 15 dias', 'Fulano')] == 1
    assert por_celula[('Entre 6 meses e 1 ano', 'Fulano')] == 1
    assert por_celula[('Há 2 anos', 'Sicrano')] == 0, 'julgado ainda contando'
    cur.connection.rollback()


@teste
def resumo_do_acervo_conta_processo_redistribuido_uma_vez(cur):
    """Duas distribuições, um processo: conta na cadeira mais recente."""
    autenticar(cur)
    cur.execute("delete from public.julgados_cj")
    cur.execute("delete from public.acervo_cj")
    cur.execute("""insert into public.acervo_cj
                     (num_processo, relator, data_distribuicao, defesa, origem) values
                   ('202600029000004', 'Antigo', current_date - 300, true, 'sorteio'),
                   ('202600029000004', 'Atual',  current_date - 3,   true, 'sorteio')""")

    cur.execute('select faixa, relator, processos from public.resumo_acervo_cj()'
                ' where processos > 0')
    assert cur.fetchall() == [('Até 15 dias', 'Atual', 1)], 'contou a distribuição antiga'
    cur.connection.rollback()


@teste
def resumo_do_acervo_conta_redistribuicao_posterior_ao_julgado(cur):
    """Julgado antigo não pode esconder a distribuição que veio depois dele.

    "Não julgado" era "não aparece em julgados_cj", sem correlação de data
    nenhuma. Enquanto julgados_cj só tinha jun/2026 em diante isso não mordia;
    com a série crescendo, um processo julgado e depois redistribuído sumiria do
    painel — distribuído e invisível.
    """
    autenticar(cur)
    cur.execute("delete from public.julgados_cj")
    cur.execute("delete from public.acervo_cj")
    cur.execute("""insert into public.acervo_cj
                     (num_processo, relator, data_distribuicao, defesa, origem) values
                   ('202600029000005', 'CJ1', current_date - 300, true, 'sorteio'),
                   ('202600029000005', 'CJ3', current_date - 3,   true, 'sorteio')""")
    cur.execute("""insert into public.julgados_cj
                     (num_processo, data_sessao, voto, status)
                   values ('202600029000005', current_date - 200, 'Manter', 'Julgado')""")

    cur.execute('select faixa, relator, processos from public.resumo_acervo_cj()'
                ' where processos > 0')
    assert cur.fetchall() == [('Até 15 dias', 'CJ3', 1)], \
        'a redistribuição posterior ao julgado tem de aparecer'

    cur.execute('select num_processo, relator from public.processos_acervo_cj()')
    assert cur.fetchall() == [('202600029000005', 'CJ3')]

    # E o julgado que veio DEPOIS da distribuição continua tirando do painel.
    cur.execute("""insert into public.julgados_cj
                     (num_processo, data_sessao, voto, status)
                   values ('202600029000005', current_date, 'Manter', 'Julgado')""")
    cur.execute('select coalesce(sum(processos), 0) from public.resumo_acervo_cj()')
    assert cur.fetchone()[0] == 0
    cur.connection.rollback()


@teste
def de_para_das_cadeiras_esta_completo(cur):
    """Cinco cadeiras, cinco conselheiros, sem ambiguidade.

    Duas cadeiras com a mesma pessoa vigente, ou a mesma cadeira com dois
    ocupantes ao mesmo tempo, quebrariam a tradução em silêncio.
    """
    cur.execute("""select cadeira, conselheiro from public.cadeiras_cj
                    where ate is null order by cadeira""")
    vigentes = cur.fetchall()
    assert len(vigentes) == 5, vigentes
    assert len({c for c, _ in vigentes}) == 5, 'cadeira repetida na mesma vigência'
    assert len({n for _, n in vigentes}) == 5, 'conselheiro em duas cadeiras'


@teste
def sorteio_da_cj_grava_cadeira_e_o_front_sabe_o_nome(cur):
    """O de-para do supabase.js espelha a tabela do banco.

    O que vai para acervo_cj é a cadeira; o nome existe no front só para o
    hover. Se as duas listas divergirem, a tela passa a anunciar o conselheiro
    errado — e nada no banco perceberia, porque o nome não é gravado.
    """
    fonte = (RAIZ / 'assets' / 'js' / 'supabase.js').read_text(encoding='utf-8')
    trecho = fonte[fonte.index('const CADEIRAS_CJ'):]
    no_front = dict(re.findall(r"(CJ\d+):\s*'([^']+)'", trecho[:trecho.index('};')]))
    assert len(no_front) == 5, no_front

    cur.execute('select cadeira, conselheiro from public.cadeiras_cj where ate is null')
    no_banco = dict(cur.fetchall())
    assert no_front == no_banco, f'front={no_front} banco={no_banco}'


@teste
def conversao_troca_nome_por_cadeira_e_poupa_o_resto(cur):
    """O UPDATE da migração 20260824180000, na fronteira que ele promete.

    Quem está no de-para vira cadeira. Quem não está — conselheiro de
    composição anterior, que o histórico de 2024 e 2025 tem — fica pelo nome:
    inventar o número da cadeira dele seria pior do que não traduzir.

    O comando é LIDO do arquivo da migração, não copiado para cá: uma cópia
    continuaria passando depois de a migração ser alterada ou quebrada.
    """
    cur.execute("""insert into public.acervo_cj
                     (num_processo, relator, data_distribuicao, defesa, origem) values
                   ('900000000000090', 'Dorivan de Souza Lima',  date '2026-07-01', true, 'sorteio'),
                   ('900000000000091', 'Paulo Otoni Ribeiro',    date '2026-07-01', false,'sorteio'),
                   ('900000000000092', 'Conselheiro De Antes',   date '2026-07-01', true, 'planilha')""")

    cur.execute(comando_da_migracao('acervo_cj'))

    cur.execute("""select num_processo, relator from public.acervo_cj
                    where num_processo in ('900000000000090','900000000000091','900000000000092')
                    order by num_processo""")
    convertidos = cur.fetchall()
    assert convertidos == [('900000000000090', 'CJ3'),
                           ('900000000000091', 'CJ1'),
                           ('900000000000092', 'Conselheiro De Antes')], convertidos
    cur.connection.rollback()


@teste
def reimportar_a_planilha_nao_duplica_o_relator(cur):
    """Rodar o SQL da importação duas vezes não cria uma segunda coluna.

    A planilha traz o NOME do conselheiro; o banco guarda a CADEIRA. Se a
    tradução ficasse para um UPDATE depois do insert, a segunda importação
    gravaria a linha pelo nome — que não colide com a da cadeira, porque a
    restrição inclui relator — e o painel passaria a ter duas colunas para a
    mesma pessoa. O SQL vem do importador de verdade, não de uma cópia.
    """
    import importar_planilha as imp

    colunas = ['num_processo', 'relator', 'data_distribuicao', 'defesa', 'origem']
    linhas = [
        {'num_processo': '900000000000093', 'relator': 'Dorivan de Souza Lima',
         'data_distribuicao': date(2026, 7, 1), 'defesa': True, 'origem': 'planilha'},
        # Composição anterior: sem cadeira no de-para, fica pelo nome.
        {'num_processo': '900000000000094', 'relator': 'Conselheiro De Antes',
         'data_distribuicao': date(2026, 7, 1), 'defesa': None, 'origem': 'planilha'},
    ]
    sql = imp.gerar_sql('acervo_cj', colunas, linhas, 'acervo_cj_distribuicao_unica',
                        'teste', 'data_distribuicao')
    cur.execute(sql)
    cur.execute(sql)

    cur.execute("""select num_processo, relator from public.acervo_cj
                    where num_processo in ('900000000000093', '900000000000094')
                    order by num_processo, relator""")
    gravado = cur.fetchall()
    assert gravado == [('900000000000093', 'CJ3'),
                       ('900000000000094', 'Conselheiro De Antes')], gravado
    cur.connection.rollback()


@teste
def cadeira_nao_aceita_dois_ocupantes_vigentes(cur):
    """Duas linhas com `ate` nulo dobrariam cada célula do painel.

    A chave primária é (cadeira, desde) e não impede isso: o de-para entra no
    FROM da mesma consulta que conta os processos, então uma cadeira com dois
    períodos abertos multiplicaria o count por dois — sem erro nenhum na tela.
    """
    try:
        cur.execute("""insert into public.cadeiras_cj (cadeira, conselheiro, desde)
                       values ('CJ3', 'Outro Conselheiro', date '2027-01-01')""")
        raise AssertionError('aceitou dois ocupantes vigentes na mesma cadeira')
    except psycopg2.errors.UniqueViolation:
        pass
    cur.connection.rollback()

    # Fechar o período anterior é o caminho previsto, e continua permitido.
    cur.execute("update public.cadeiras_cj set ate = date '2026-12-31' where cadeira = 'CJ3'")
    cur.execute("""insert into public.cadeiras_cj (cadeira, conselheiro, desde)
                   values ('CJ3', 'Outro Conselheiro', date '2027-01-01')""")
    cur.connection.rollback()


@teste
def detalhe_do_painel_confere_com_a_contagem(cur):
    """A lista que o card abre tem exatamente o número que o bloco mostrava.

    São duas funções com a mesma definição de pendente e as mesmas faixas. Se
    uma mudar sem a outra, o painel diz 22 e o card abre 21 — e nada mais no
    sistema perceberia.

    Sem a planilha (CI), acervo_cj está vazio e a comparação viraria 0 == 0: o
    teste semeia faixas e cadeiras diferentes, mais um julgado e um
    redistribuído, que são justamente os pontos onde as duas podem divergir.
    """
    autenticar(cur)
    cur.execute("delete from public.julgados_cj")
    cur.execute("delete from public.acervo_cj")
    cur.execute("""insert into public.acervo_cj
                     (num_processo, relator, data_distribuicao, defesa, origem) values
                   ('202600029000011', 'CJ1', current_date - 5,   true, 'sorteio'),
                   ('202600029000012', 'CJ1', current_date - 200, true, 'sorteio'),
                   ('202600029000013', 'CJ2', current_date - 900, true, 'sorteio'),
                   ('202600029000014', 'CJ2', current_date - 40,  true, 'sorteio'),
                   ('202600029000015', 'CJ3', current_date - 60,  true, 'sorteio'),
                   ('202600029000016', 'CJ1', current_date - 300, true, 'sorteio'),
                   ('202600029000016', 'CJ3', current_date - 3,   true, 'sorteio')""")
    # julgado sai das duas
    cur.execute("""insert into public.julgados_cj (num_processo, data_sessao, pauta)
                   values ('202600029000015', current_date - 10, 1)""")

    cur.execute('select sum(processos)::int from public.resumo_acervo_cj()')
    do_painel = cur.fetchone()[0]
    cur.execute('select count(*) from public.processos_acervo_cj()')
    do_detalhe = cur.fetchone()[0]
    assert do_painel == do_detalhe, f'painel={do_painel} detalhe={do_detalhe}'

    # e célula a célula, não só no total
    cur.execute('select ordem, relator, processos from public.resumo_acervo_cj()'
                ' where processos > 0')
    for ordem, relator, esperado in cur.fetchall():
        cur.execute('select count(*) from public.processos_acervo_cj(%s, %s)', (ordem, relator))
        achado = cur.fetchone()[0]
        assert achado == esperado, f'faixa {ordem} / {relator}: painel={esperado} detalhe={achado}'
    assert do_painel == 5, f'semeadura não exercita as duas funções: {do_painel}'
    cur.connection.rollback()


@teste
def detalhe_do_painel_exige_sessao(cur):
    """Ela lê acervo_cj processo a processo — sem a guarda, seria um vazamento."""
    cur.execute("select set_config('request.jwt.claims', '', true)")
    try:
        cur.execute('select * from public.processos_acervo_cj()')
        raise AssertionError('respondeu sem sessão autenticada')
    except psycopg2.errors.InvalidAuthorizationSpecification:
        pass
    cur.connection.rollback()


@teste
def detalhe_do_painel_so_e_executavel_por_authenticated(cur):
    cur.execute("""select coalesce(string_agg(distinct grantee, ',' order by grantee), '')
                     from information_schema.role_routine_grants
                    where routine_schema = 'public'
                      and routine_name = 'processos_acervo_cj'
                      and grantee in ('anon', 'authenticated', 'service_role', 'PUBLIC')""")
    assert cur.fetchone()[0] == 'authenticated'


@teste
def detalhe_lista_o_processo_uma_vez_so(cur):
    """Redistribuído aparece uma vez, na cadeira e na data mais recentes."""
    autenticar(cur)
    cur.execute("delete from public.julgados_cj")
    cur.execute("delete from public.acervo_cj")
    cur.execute("""insert into public.acervo_cj
                     (num_processo, relator, data_distribuicao, defesa, origem) values
                   ('900000000000070', 'CJ1', current_date - 300, true, 'sorteio'),
                   ('900000000000070', 'CJ3', current_date - 5,   true, 'sorteio')""")
    cur.execute('select num_processo, relator, dias from public.processos_acervo_cj()')
    assert cur.fetchall() == [('900000000000070', 'CJ3', 5)], 'contou a distribuição antiga'
    cur.connection.rollback()


@teste
def painel_traduz_cadeira_e_deixa_o_resto_intacto(cur):
    """Cadeira vem com o nome para o hover; sem de-para, o rótulo se repete.

    O coalesce da função é o que evita hover vazio: uma coluna cujo relator não
    é cadeira mostra o próprio valor, em vez de um title em branco. Sem a
    planilha (CI), acervo_cj está vazio — o próprio teste tem que semear os
    dois casos, senão "painel vazio" nem chega a exercitar o coalesce.
    """
    autenticar(cur)
    cur.execute("delete from public.julgados_cj")
    cur.execute("delete from public.acervo_cj")
    cur.execute("""insert into public.acervo_cj
                     (num_processo, relator, data_distribuicao, defesa, origem) values
                   ('202600029000005', 'CJ2',                    current_date - 5, true, 'sorteio'),
                   ('202600029000006', 'Conselheiro Sem Cadeira', current_date - 5, true, 'sorteio')""")

    cur.execute('select distinct relator, conselheiro from public.resumo_acervo_cj()')
    pares = dict(cur.fetchall())
    assert pares, 'painel vazio'

    cur.execute('select cadeira, conselheiro from public.cadeiras_cj where ate is null')
    depara = dict(cur.fetchall())
    for relator, conselheiro in pares.items():
        esperado = depara.get(relator, relator)
        assert conselheiro == esperado, f'{relator} -> {conselheiro}, esperado {esperado}'
    assert pares['CJ2'] != 'CJ2', 'CJ2 deveria mostrar o conselheiro, não a própria cadeira'
    assert pares['Conselheiro Sem Cadeira'] == 'Conselheiro Sem Cadeira', 'sem de-para deveria repetir o rótulo'
    cur.connection.rollback()


@teste
def creg_grava_no_proprio_acervo(cur):
    """Desde 27/08/2026 o sorteio do Conselho vai para acervo_creg.

    É a única tabela em que o CREG grava: a antiga saiu do schema, o front
    aponta para o acervo e é de lá que os julgados do Conselho saem.
    """
    cur.execute("""insert into acervo_creg
                   (num_processo, unidade, data_distribuicao, assunto, recurso)
                   values ('202600029000999', 'CREG3', current_date,
                           'Requerimento', 'Não se aplica')
                   returning unidade""")
    assert cur.fetchone()[0] == 'CREG3'
    cur.connection.rollback()


@teste
@exige_planilha
def dados_importados_nao_produzem_nenhum_erro(cur):
    """Roda o verificacao_cj.sql contra os dados reais e exige zero ERRO.

    Os AVISOs são qualidade de dado herdada da planilha e estão documentados em
    FLUXO-CJ.md; ERRO é quebra de regra do sistema e não pode existir.
    """
    cur.execute((RAIZ / 'sql' / 'verificacao_cj.sql').read_text(encoding='utf-8'))
    achados = cur.fetchall()
    assert achados, 'a verificação não devolveu nenhuma linha'

    erros = [(nome, qtd) for grav, nome, qtd, _ in achados if grav == 'ERRO']
    assert not erros, erros


@teste
def rotulos_da_pagina_batem_com_os_do_banco(cur):
    """julgados.js e registrar_votos precisam aceitar exatamente a mesma lista.

    Se divergirem, a secretaria escolhe um rótulo no seletor e o banco recusa na
    hora de salvar — falha que só apareceria em produção.
    """
    pagina = (RAIZ / 'assets' / 'js' / 'julgados.js').read_text(encoding='utf-8')
    schema = (RAIZ / 'sql' / 'schema.sql').read_text(encoding='utf-8')

    # São duas funções de registro no schema, uma por colegiado, e as duas
    # trazem as mesmas palavras. O recorte pelo corpo da função é o que garante
    # que a página da Câmara está sendo medida contra a lista da Câmara.
    corpo = re.search(r'function public\.registrar_votos\(itens jsonb\).*?\n\$\$;',
                      schema, re.S)
    assert corpo, 'registrar_votos não encontrada no schema'
    corpo = corpo.group(0)

    def rotulos(fonte, padrao):
        achado = re.search(padrao, fonte, re.S)
        assert achado, padrao
        return re.findall(r"'([^']+)'", achado.group(1))

    assert rotulos(pagina, r'const VOTOS = \[([^\]]+)\]') == \
           rotulos(corpo, r"'voto', ''\)\s*not in \(([^)]+)\)")
    assert rotulos(pagina, r'const STATUS = \[([^\]]+)\]') == \
           rotulos(corpo, r"'status', ''\)\s*not in \(([^)]+)\)")


@teste
def colunas_que_o_index_js_envia_existem(cur):
    """As chaves do payload de index.js batem com as colunas de cada tabela."""
    fonte = (RAIZ / 'assets' / 'js' / 'index.js').read_text(encoding='utf-8')
    trecho = fonte[fonte.index('function linhasParaBanco'):fonte.index('async function salvar')]
    enviadas = {linha.split(':')[0].strip()
                for linha in trecho.splitlines()
                if ':' in linha and linha.strip()[0].isalpha() and '//' not in linha}

    cur.execute("""select table_name, column_name from information_schema.columns
                    where table_schema = 'public'
                      and table_name in ('acervo_cj', 'acervo_creg')""")
    existentes = {c for _, c in cur.fetchall()}
    assert enviadas and enviadas <= existentes, sorted(enviadas - existentes)

    assert "TABELAS = { CJ: 'acervo_cj', CREG: 'acervo_creg' }" in fonte

    # Na CJ a 6ª coluna é Defesa, não Recurso: defesa sai no ramo do CJ (o
    # primeiro) e recurso só no do CREG, uma vez em cada.
    assert trecho.count('defesa:') == 1 and trecho.count('recurso:') == 1
    assert trecho.index("=== 'CJ'") < trecho.index('defesa:') < trecho.index('recurso:')


@teste
@exige_planilha
def backup_e_restauracao_fecham_o_ciclo(cur):
    """Backup → estrago → restauração devolve o banco exatamente como estava.

    Testa o caminho inteiro porque um backup que não se restaura não é backup:
    as colunas id são identity e dias_dt/periodo_dt são geradas, e as duas
    coisas quebram um `insert ... select *` ingênuo.

    Os scripts abrem transação própria em outra conexão, então este teste
    fecha a sua antes de cada um — senão os locks se cruzam e o banco trava.
    """
    def conta():
        cur.connection.commit()
        with PG.conectar() as c, c.cursor() as k:
            return (uma(k, 'select count(*) from acervo_cj'),
                    uma(k, 'select count(*) from julgados_cj'))

    def valor(sql):
        with PG.conectar() as c, c.cursor() as k:
            return uma(k, sql)

    antes = conta()

    PG.rodar_arquivo(RAIZ / 'sql' / 'backup_cj.sql')

    # Um estrago do tamanho do que a limpeza fez em produção: julgados zerados e
    # o acervo reduzido ao que nunca foi julgado.
    PG.executar("""
        delete from public.julgados_cj;
        delete from public.acervo_cj a
         where exists (select 1 from backup_cj.julgados_cj j
                        where j.num_processo = a.num_processo);
    """)

    acervo, julgados = conta()
    assert julgados == 0 and 0 < acervo < antes[0]

    PG.rodar_arquivo(RAIZ / 'sql' / 'restaurar_cj.sql')

    assert conta() == antes, 'a restauração não devolveu os mesmos totais'
    assert valor('select count(*) from julgados_cj where dias_dt is null') == 0
    assert valor("""select count(*) from julgados_cj j
                     where j.acervo_id is not null
                       and not exists (select 1 from acervo_cj a where a.id = j.acervo_id)""") == 0

    PG.executar('drop schema backup_cj cascade')


@teste
def migracoes_estao_todas_cobertas(cur):
    """Nenhuma migração pode entrar no diretório sem ser executada por aqui.

    As duas do Conselho — a que cria as tabelas, gatilhos, RPCs e RLS dele e a
    que copia o sorteio de 27/08 — ficaram meses sem teste nenhum justamente
    porque a lista era escrita à mão.
    """
    arquivos = {a.name for a in (RAIZ / 'supabase' / 'migrations').glob('*.sql')}
    aplicadas = {a.name for a in migracoes_a_aplicar()} | {MIGRACAO.name}

    assert arquivos - aplicadas == MIGRACOES_SUPERADAS, (
        'migração fora da bateria: ' + ', '.join(sorted(arquivos - aplicadas)))
    assert MIGRACOES_SUPERADAS <= arquivos, 'MIGRACOES_SUPERADAS cita arquivo que não existe'
    for constante in (MIGRACAO_CADEIRAS, MIGRACAO_HARDENING, MIGRACAO_INTERESSADO,
                      MIGRACAO_REMOCAO):
        assert constante.name in aplicadas, constante.name


@teste
def schema_declara_toda_funcao_que_existe_no_banco(cur):
    """O schema.sql tem de conhecer toda função que o banco tem.

    migracoes_reproduzem_o_schema pega a função que o schema MUDA, mas não a
    que ele esquece: se uma migração cria uma função que o schema.sql nunca
    menciona, reaplicar o schema não mexe nela, `antes` e `depois` ficam iguais
    e a comparação passa verde.

    Foi por esse buraco que historico_sorteios, processos_sorteio e
    historico_marco viveram em produção sem estar no repositório — aplicadas
    direto no banco, sem commit, e nenhum teste tinha como notar. A conta aqui
    é a inversa: o conjunto de funções do banco tem de ser exatamente o que o
    arquivo declara, nem mais nem menos.

    Roda ANTES de migracoes_reproduzem_o_schema de propósito: neste ponto o
    banco veio só das migrações, que é o estado que produção tem.
    """
    fonte = (RAIZ / 'sql' / 'schema.sql').read_text(encoding='utf-8')
    declaradas = set(re.findall(
        r'create\s+(?:or\s+replace\s+)?function\s+public\.(\w+)', fonte))

    cur.execute("""select distinct p.proname
                     from pg_proc p
                     join pg_namespace n on n.oid = p.pronamespace
                    where n.nspname = 'public' and p.prokind = 'f'""")
    no_banco = {nome for (nome,) in cur.fetchall()}

    assert no_banco == declaradas, (
        'no banco e fora do schema.sql: '
        + (', '.join(sorted(no_banco - declaradas)) or '—')
        + ' / no schema.sql e fora do banco: '
        + (', '.join(sorted(declaradas - no_banco)) or '—'))


@teste
def migracoes_reproduzem_o_schema(cur):
    """Produção aplica migrações; sql/schema.sql é só o estado desejado.

    Este banco veio do schema.sql, foi rebaixado ao estado anterior por
    preparar_upgrade_da_migracao e reconstruído SÓ pelo diretório de migrações.
    Reaplicar o schema.sql agora não pode mudar função nenhuma: se mudar, é
    porque alguma migração não carregou o que o schema já tem, e produção ficaria
    com a versão velha para sempre — sem nada ficar vermelho.
    """
    def definicoes():
        cur.execute("""
            select p.proname || '(' ||
                   pg_get_function_identity_arguments(p.oid) || ')',
                   pg_get_functiondef(p.oid)
              from pg_proc p
              join pg_namespace n on n.oid = p.pronamespace
             where n.nspname = 'public'
             order by 1""")
        return dict(cur.fetchall())

    antes = definicoes()
    assert antes, 'nenhuma função em public: o cenário do teste não montou'
    cur.connection.commit()

    rodar_arquivo(RAIZ / 'sql' / 'schema.sql')
    depois = definicoes()

    divergentes = sorted(set(antes) ^ set(depois)) or sorted(
        nome for nome in antes if antes[nome] != depois[nome])
    assert not divergentes, (
        'o schema.sql tem versão diferente da que as migrações entregam: '
        + ', '.join(divergentes))


# ── Runner ───────────────────────────────────────────────────────────────────

# Migração que uma posterior substituiu: 20260824180000 derruba e recria
# resumo_acervo_cj com a coluna `conselheiro`, então rodar a versão de
# 20260824121500 sobre o schema atual falha com "cannot change return type" —
# pelo motivo certo. Em produção ela rodou na ordem, sobre o banco de antes.
MIGRACOES_SUPERADAS = {'20260824121500_painel_acervo_cj.sql'}


def migracoes_a_aplicar():
    """As migrações posteriores a MIGRACAO, em ordem cronológica."""
    return [a for a in sorted((RAIZ / 'supabase' / 'migrations').glob('*.sql'))
            if a.name > MIGRACAO.name and a.name not in MIGRACOES_SUPERADAS]


def preparar_upgrade_da_migracao():
    """Volta só os deltas desta migração ao estado do schema no HEAD."""
    PG.executar("""
        -- A tabela do sorteio antigo do Conselho, recriada no formato que ela
        -- tinha antes desta migração: sem a restrição de 15 dígitos, sem o
        -- índice de distribuição única e sem a coluna interessado, que só
        -- chegaria em 20260827160000. O schema.sql não a cria mais — quem a
        -- derrubou foi a migração 20260902120000 —, então é aqui que ela
        -- volta, e só para que as migrações que mexem nela tenham em que
        -- mexer. No fim da bateria ela é derrubada de novo, pela migração de
        -- verdade.
        create table public.processos_sorteados (
          id                bigint generated always as identity primary key,
          criado_em         timestamptz not null default now(),
          modo              text        not null check (modo = 'CREG'),
          data_hora         timestamptz not null,
          ordem             int         not null,
          num_processo      text        not null,
          assunto           text        not null,
          data_distribuicao date        not null,
          recurso           text        not null,
          unidade           text        not null
        );
        create index idx_processos_sorteados_modo_data
          on public.processos_sorteados (modo, data_hora desc);
        alter table public.processos_sorteados enable row level security;
        create policy "usuario autenticado pode inserir"
          on public.processos_sorteados for insert to authenticated with check (true);

        delete from public.pautas_cj where url = 'marco:inicio-da-serie';

        create or replace function public.julgados_cj_derivar_do_acervo()
        returns trigger
        language plpgsql
        security definer
        set search_path = ''
        as $$
        declare
          origem public.acervo_cj%rowtype;
        begin
          if new.data_distribuicao is not null then
            select * into origem
              from public.acervo_cj
             where num_processo = new.num_processo
               and data_distribuicao = new.data_distribuicao
             order by id
             limit 1;
          end if;

          if origem.id is null then
            select * into origem
              from public.acervo_cj
             where num_processo = new.num_processo
               and data_distribuicao <= new.data_sessao
             order by data_distribuicao desc, id desc
             limit 1;
          end if;

          if origem.id is null then
            select * into origem
              from public.acervo_cj
             where num_processo = new.num_processo
             order by data_distribuicao, id
             limit 1;
          end if;

          new.acervo_id         := origem.id;
          new.relator           := coalesce(new.relator, origem.relator);
          new.defesa            := coalesce(new.defesa, origem.defesa);
          new.data_distribuicao := coalesce(new.data_distribuicao, origem.data_distribuicao);
          return new;
        end;
        $$;

        create or replace function public.auth_email()
        returns text
        language sql
        stable
        as $$
          select nullif(current_setting('request.jwt.claims', true), '')::jsonb ->> 'email'
        $$;
        alter function public.auth_email() reset all;

        create or replace function public.registrar_votos(itens jsonb)
        returns int
        language plpgsql
        security definer
        set search_path = ''
        as $$
        declare
          quem     text := coalesce(public.auth_email(), 'desconhecido');
          invalido int;
          gravados int;
        begin
          if jsonb_typeof(itens) is distinct from 'array' then
            raise exception 'registrar_votos espera uma lista de itens';
          end if;

          select count(*) into invalido
            from jsonb_array_elements(itens) i
           where coalesce(i ->> 'id', '') !~ '^[0-9]+$'
              or nullif(i ->> 'voto', '')   not in ('Manter', 'Anular', 'Vista')
              or nullif(i ->> 'status', '') not in ('Julgado', 'Retornou', 'Retirado', 'Vista');

          if invalido > 0 then
            raise exception 'id, voto ou status fora do permitido (% item(ns))', invalido;
          end if;

          update public.julgados_cj j
             set voto           = nullif(i ->> 'voto', ''),
                 status         = nullif(i ->> 'status', ''),
                 atualizado_em  = now(),
                 atualizado_por = quem
            from jsonb_array_elements(itens) i
           where j.id = (i ->> 'id')::bigint
             and (j.voto is null or j.status is null or j.atualizado_em is not null);

          get diagnostics gravados = row_count;
          return gravados;
        end;
        $$;

        grant all on public.processos_sorteados, public.acervo_cj,
                     public.julgados_cj, public.pautas_cj
          to anon, authenticated;
        -- Nomeadas uma a uma, e não "all sequences in schema public": as
        -- sequências do Conselho Regulador nasceram depois desta migração e
        -- não podem entrar no estado que ela encontra.
        grant all on sequence public.processos_sorteados_id_seq,
                              public.acervo_cj_id_seq,
                              public.julgados_cj_id_seq,
                              public.pautas_cj_id_seq
          to anon, authenticated;
        grant execute on function public.julgados_cj_derivar_do_acervo()
          to public, anon, authenticated;
        grant execute on function public.auth_email() to public, anon, authenticated;
        grant execute on function public.registrar_votos(jsonb)
          to anon, service_role;

        -- data_hora dentro de 27/08/2026 de propósito: é a janela que a
        -- migração 20260828090000 copia para acervo_creg. Com now() a cópia
        -- não alcançava linha nenhuma e passava verde sem ter feito nada —
        -- migracao_levou_o_sorteio_antigo_para_o_acervo mede o resultado dela.
        insert into public.processos_sorteados
          (modo, data_hora, ordem, num_processo, assunto,
           data_distribuicao, recurso, unidade)
        values
          ('CREG', timestamptz '2026-08-27 11:07:26-03', 1, '123421', 'Requerimento',
           date '2026-08-20', 'Com recurso', 'CREG2'),
          ('CREG', timestamptz '2026-08-27 11:07:26-03', 2, '1234', 'Requerimento',
           date '2026-08-20', 'Sem recurso', 'CREG3'),
          ('CREG', timestamptz '2026-08-27 11:07:26-03', 3, '202600029000777', 'Requerimento',
           date '2026-08-20', 'Com recurso', 'CREG2');
    """)


def preparar_banco(planilha):
    rodar_arquivo(RAIZ / 'sql' / 'schema.sql')
    preparar_upgrade_da_migracao()
    rodar_arquivo(MIGRACAO)

    if planilha:
        subprocess.run([sys.executable, str(RAIZ / 'dados' / 'importar_planilha.py'),
                        str(planilha)], check=True, capture_output=True)
        rodar_arquivo(RAIZ / 'dados' / 'acervo_cj.sql')
        rodar_arquivo(RAIZ / 'dados' / 'julgados_cj.sql')

    # Depois da planilha, e sempre: é a única forma de o arquivo da migração ser
    # de fato executado em CI. Ele tem de ser repetível — rodar sobre um banco
    # que já veio do schema.sql, com o dado já em cadeira, não pode falhar nem
    # converter nada duas vezes.
    # Todas as migrações posteriores, em ordem de nome — que é a ordem
    # cronológica. Produção aplica o diretório inteiro, e migração que nenhum
    # teste executa chega lá sem nunca ter rodado; migracoes_estao_todas_cobertas
    # não deixa uma nova passar batida. Depois da planilha porque
    # MIGRACAO_CADEIRAS converte o dado importado, e todas têm de ser repetíveis
    # sobre um banco que já veio do schema.sql.
    #
    # O schema.sql NÃO é reaplicado aqui. Reaplicá-lo refazia sozinho tudo o que
    # preparar_upgrade_da_migracao havia rebaixado, e toda assertiva sobre
    # migração passava independentemente do conteúdo dela. Quem fecha o cerco no
    # lugar dele é migracoes_reproduzem_o_schema, no fim desta bateria.
    for arquivo in migracoes_a_aplicar():
        rodar_arquivo(arquivo)


def main(argv):
    global PLANILHA
    caminho = Path(argv[1]) if len(argv) > 1 else PLANILHA_PADRAO
    if not caminho.is_file():
        print(f'AVISO: planilha não encontrada em {caminho} — testes dela serão pulados.\n')
        caminho = None

    PG.subir()
    try:
        preparar_banco(caminho)
        if caminho:
            PLANILHA = Planilha(caminho)

        falhas = executados = pulados = 0
        with PG.conectar() as conn:
            for fn in testes:
                if getattr(fn, 'precisa_planilha', False) and not caminho:
                    print(f'PULA  {fn.__name__}')
                    pulados += 1
                    continue
                executados += 1
                with conn.cursor() as cur:
                    try:
                        fn(cur)
                        conn.commit()
                        print(f'ok    {fn.__name__}')
                    except Exception as e:
                        conn.rollback()
                        falhas += 1
                        print(f'FALHA {fn.__name__}: {type(e).__name__}: {e}')

        print(f'\n{executados - falhas}/{executados} testes passaram; {pulados} pulados.')
        return 1 if falhas else 0
    finally:
        PG.derrubar()


if __name__ == '__main__':
    sys.exit(main(sys.argv))
